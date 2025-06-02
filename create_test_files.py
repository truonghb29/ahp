#!/usr/bin/env python3
"""
Script test để tạo file Excel mẫu cho chức năng import
"""
from openpyxl import Workbook
import numpy as np

def create_sample_candidates():
    """Tạo file Excel mẫu cho danh sách ứng viên"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Danh sach ung vien"
    
    # Header
    ws['A1'] = "Tên ứng viên"
    
    # Dữ liệu mẫu
    candidates = [
        "Nguyễn Văn An",
        "Trần Thị Bình", 
        "Lê Văn Cường",
        "Phạm Thị Dung",
        "Hoàng Văn Em"
    ]
    
    for i, candidate in enumerate(candidates, start=2):
        ws[f'A{i}'] = candidate
    
    wb.save('test_candidates.xlsx')
    print(f"Đã tạo file test_candidates.xlsx với {len(candidates)} ứng viên")

def create_sample_criteria_matrix():
    """Tạo file Excel mẫu cho ma trận tiêu chí 4x4"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ma tran tieu chi"
    
    # Ma trận 4x4 mẫu theo AHP
    matrix = [
        [1, 3, 5, 7],
        [1/3, 1, 3, 5],
        [1/5, 1/3, 1, 3],
        [1/7, 1/5, 1/3, 1]
    ]
    
    for i, row in enumerate(matrix):
        for j, value in enumerate(row):
            ws.cell(row=i+1, column=j+1, value=value)
    
    wb.save('test_criteria_matrix_4x4.xlsx')
    print("Đã tạo file test_criteria_matrix_4x4.xlsx")

def create_sample_candidate_matrix():
    """Tạo file Excel mẫu cho ma trận ứng viên 5x5"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ma tran ung vien"
    
    # Ma trận 5x5 mẫu
    matrix = [
        [1, 2, 3, 4, 5],
        [1/2, 1, 2, 3, 4],
        [1/3, 1/2, 1, 2, 3],
        [1/4, 1/3, 1/2, 1, 2],
        [1/5, 1/4, 1/3, 1/2, 1]
    ]
    
    for i, row in enumerate(matrix):
        for j, value in enumerate(row):
            ws.cell(row=i+1, column=j+1, value=value)
    
    wb.save('test_candidate_matrix_5x5.xlsx')
    print("Đã tạo file test_candidate_matrix_5x5.xlsx")

if __name__ == "__main__":
    create_sample_candidates()
    create_sample_criteria_matrix()
    create_sample_candidate_matrix()
    print("Hoàn thành tạo các file Excel mẫu!")
