#!/usr/bin/env python3
"""
Tạo ma trận tiêu chí 6x6 với tên tiêu chí để test import
"""
from openpyxl import Workbook

def create_criteria_matrix_6x6():
    """Tạo ma trận so sánh tiêu chí 6x6 với tên tiêu chí"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ma tran tieu chi 6x6"
    
    # Danh sách tên tiêu chí
    criteria_names = [
        "Kinh nghiệm",
        "Thái độ",
        "Kỹ năng giao tiếp", 
        "Khả năng học hỏi",
        "Làm việc nhóm",
        "Sáng tạo"
    ]
    
    # Ma trận 6x6 theo thang đo Saaty
    matrix = [
        [1,     3,     2,     4,     2,     3],     # Kinh nghiệm
        [1/3,   1,     1/2,   2,     1/3,   1],     # Thái độ  
        [1/2,   2,     1,     3,     1/2,   2],     # Kỹ năng giao tiếp
        [1/4,   1/2,   1/3,   1,     1/4,   1/2],   # Khả năng học hỏi
        [1/2,   3,     2,     4,     1,     3],     # Làm việc nhóm
        [1/3,   1,     1/2,   2,     1/3,   1]      # Sáng tạo
    ]
    
    # Thêm header với tên tiêu chí (hàng 1, bắt đầu từ cột 2)
    for j, criterion in enumerate(criteria_names):
        ws.cell(row=1, column=j+2, value=criterion)
    
    # Thêm tên tiêu chí ở cột đầu và ma trận
    for i, (criterion, row_data) in enumerate(zip(criteria_names, matrix)):
        ws.cell(row=i+2, column=1, value=criterion)  # Tên tiêu chí ở cột đầu
        for j, value in enumerate(row_data):
            ws.cell(row=i+2, column=j+2, value=value)  # Ma trận bắt đầu từ cột 2
    
    # Lưu file
    filename = 'test_criteria_matrix_6x6.xlsx'
    wb.save(filename)
    print(f"Đã tạo file {filename}")
    wb.close()
    
    return filename

def create_candidate_matrices_6_criteria():
    """Tạo 6 ma trận ứng viên 3x3 cho 6 tiêu chí"""
    
    # Tên ứng viên
    candidates = ["Ứng viên A", "Ứng viên B", "Ứng viên C"]
    
    # Ma trận cho từng tiêu chí với chiến lược khác nhau
    matrices_data = {
        "kinh_nghiem": [
            [1, 3, 4],      # A mạnh nhất
            [1/3, 1, 2],
            [1/4, 1/2, 1]
        ],
        "thai_do": [
            [1, 1/2, 2],    # B mạnh nhất
            [2, 1, 3],
            [1/2, 1/3, 1]
        ],
        "giao_tiep": [
            [1, 1/3, 1/2],  # C mạnh nhất
            [3, 1, 2],
            [2, 1/2, 1]
        ],
        "hoc_hoi": [
            [1, 4, 3],      # A mạnh nhất
            [1/4, 1, 1/2],
            [1/3, 2, 1]
        ],
        "lam_viec_nhom": [
            [1, 1/2, 1/3],  # C mạnh nhất
            [2, 1, 1/2],
            [3, 2, 1]
        ],
        "sang_tao": [
            [1, 2, 1/2],    # Cân bằng
            [1/2, 1, 1/3],
            [2, 3, 1]
        ]
    }
    
    created_files = []
    
    for criterion_key, matrix_data in matrices_data.items():
        wb = Workbook()
        ws = wb.active
        ws.title = f"Ma tran ung vien {criterion_key}"
        
        # Thêm header với tên ứng viên (hàng 1, bắt đầu từ cột 2)
        for j, candidate in enumerate(candidates):
            ws.cell(row=1, column=j+2, value=candidate)
        
        # Thêm tên ứng viên ở cột đầu và ma trận
        for i, (candidate, row_data) in enumerate(zip(candidates, matrix_data)):
            ws.cell(row=i+2, column=1, value=candidate)  # Tên ứng viên ở cột đầu
            for j, value in enumerate(row_data):
                ws.cell(row=i+2, column=j+2, value=value)  # Ma trận bắt đầu từ cột 2
        
        filename = f'test_candidate_matrix_3x3_{criterion_key}.xlsx'
        wb.save(filename)
        print(f"Đã tạo file {filename}")
        wb.close()
        created_files.append(filename)
    
    return created_files

def main():
    """Tạo tất cả các file test"""
    print("🔧 Tạo ma trận tiêu chí 6x6...")
    criteria_file = create_criteria_matrix_6x6()
    
    print("\n🔧 Tạo 6 ma trận ứng viên cho các tiêu chí...")
    candidate_files = create_candidate_matrices_6_criteria()
    
    print(f"\n✅ Đã tạo thành công tất cả các file test!")
    print(f"\n📋 File ma trận tiêu chí:")
    print(f"   - {criteria_file}")
    print(f"\n📋 File ma trận ứng viên:")
    for file in candidate_files:
        print(f"   - {file}")
    
    print(f"\n💡 Hướng dẫn sử dụng:")
    print(f"   1. Import {criteria_file} ở bước thiết lập tiêu chí")
    print(f"   2. Import từng file ma trận ứng viên tương ứng với từng tiêu chí")
    print(f"   3. Hệ thống sẽ tự động đọc tên tiêu chí và ứng viên từ file")

if __name__ == "__main__":
    main()
