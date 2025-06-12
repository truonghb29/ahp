#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script để tạo dữ liệu test Excel cho hệ thống AHP
- File danh sách 3 ứng viên
- File ma trận tiêu chí 5x5
"""

from openpyxl import Workbook
import os

def create_candidates_3_excel():
    """Tạo file Excel với 3 ứng viên"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Danh sach 3 ung vien"
    
    # Header
    ws['A1'] = "Tên ứng viên"
    
    # Dữ liệu 3 ứng viên
    candidates = [
        "Nguyễn Văn An",
        "Trần Thị Bình", 
        "Lê Minh Châu"
    ]
    
    for i, candidate in enumerate(candidates, start=2):
        ws[f'A{i}'] = candidate
    
    # Điều chỉnh độ rộng cột
    ws.column_dimensions['A'].width = 25
    
    return wb

def create_criteria_matrix_5x5_excel():
    """Tạo file Excel ma trận tiêu chí 5x5 với dữ liệu mẫu"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ma tran tieu chi 5x5"
    
    # Ma trận 5x5 với dữ liệu mẫu (theo thang đo Saaty)
    # Tiêu chí: 1-Kinh nghiệm, 2-Kỹ năng, 3-Học vấn, 4-Giao tiếp, 5-Thái độ
    matrix_data = [
        [1,     2,     3,     1/2,   2],      # Kinh nghiệm
        [1/2,   1,     2,     1/3,   1],      # Kỹ năng  
        [1/3,   1/2,   1,     1/4,   1/2],    # Học vấn
        [2,     3,     4,     1,     3],      # Giao tiếp
        [1/2,   1,     2,     1/3,   1]       # Thái độ
    ]
    
    # Ghi ma trận vào Excel
    for i in range(5):
        for j in range(5):
            # Ghi giá trị vào cell (i+1, j+1) vì Excel bắt đầu từ 1
            ws.cell(row=i+1, column=j+1, value=matrix_data[i][j])
    
    # Thêm nhãn cho các tiêu chí (tùy chọn - có thể comment out nếu không cần)
    criteria_names = ["Kinh nghiệm", "Kỹ năng", "Học vấn", "Giao tiếp", "Thái độ"]
    
    # Thêm sheet thứ 2 với tên tiêu chí để tham khảo
    ws_names = wb.create_sheet("Ten tieu chi")
    ws_names['A1'] = "STT"
    ws_names['B1'] = "Tên tiêu chí"
    
    for i, name in enumerate(criteria_names, start=2):
        ws_names[f'A{i}'] = i-1
        ws_names[f'B{i}'] = name
    
    ws_names.column_dimensions['B'].width = 20
    
    return wb

def create_candidate_matrix_3x3_sample():
    """Tạo file Excel ma trận ứng viên 3x3 mẫu cho tiêu chí đầu tiên"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ma tran ung vien 3x3"
    
    # Ma trận 3x3 cho 3 ứng viên theo tiêu chí "Kinh nghiệm"
    matrix_data = [
        [1,     2,     1/2],    # Nguyễn Văn An
        [1/2,   1,     1/3],    # Trần Thị Bình
        [2,     3,     1]       # Lê Minh Châu
    ]
    
    # Ghi ma trận vào Excel
    for i in range(3):
        for j in range(3):
            ws.cell(row=i+1, column=j+1, value=matrix_data[i][j])
    
    # Thêm sheet với tên ứng viên để tham khảo
    ws_names = wb.create_sheet("Ten ung vien")
    ws_names['A1'] = "STT"
    ws_names['B1'] = "Tên ứng viên"
    
    candidate_names = ["Nguyễn Văn An", "Trần Thị Bình", "Lê Minh Châu"]
    for i, name in enumerate(candidate_names, start=2):
        ws_names[f'A{i}'] = i-1
        ws_names[f'B{i}'] = name
    
    ws_names.column_dimensions['B'].width = 20
    
    return wb

def main():
    """Hàm chính để tạo tất cả file test"""
    print("Đang tạo các file Excel test...")
    
    # Tạo thư mục uploads nếu chưa có
    uploads_dir = "uploads"
    if not os.path.exists(uploads_dir):
        os.makedirs(uploads_dir)
    
    # 1. Tạo file danh sách 3 ứng viên
    wb_candidates = create_candidates_3_excel()
    candidates_file = os.path.join(uploads_dir, "test_candidates_3.xlsx")
    wb_candidates.save(candidates_file)
    print(f"✓ Đã tạo file: {candidates_file}")
    
    # 2. Tạo file ma trận tiêu chí 5x5
    wb_criteria = create_criteria_matrix_5x5_excel()
    criteria_file = os.path.join(uploads_dir, "test_criteria_matrix_5x5.xlsx")
    wb_criteria.save(criteria_file)
    print(f"✓ Đã tạo file: {criteria_file}")
    
    # 3. Tạo file ma trận ứng viên 3x3 mẫu
    wb_candidate_matrix = create_candidate_matrix_3x3_sample()
    candidate_matrix_file = os.path.join(uploads_dir, "test_candidate_matrix_3x3_kinh_nghiem.xlsx")
    wb_candidate_matrix.save(candidate_matrix_file)
    print(f"✓ Đã tạo file: {candidate_matrix_file}")
    
    print("\n" + "="*50)
    print("HƯỚNG DẪN SỬ DỤNG:")
    print("="*50)
    print("1. File danh sách ứng viên:")
    print(f"   - {candidates_file}")
    print("   - Chứa 3 ứng viên: Nguyễn Văn An, Trần Thị Bình, Lê Minh Châu")
    print()
    print("2. File ma trận tiêu chí 5x5:")
    print(f"   - {criteria_file}")
    print("   - Ma trận so sánh 5 tiêu chí: Kinh nghiệm, Kỹ năng, Học vấn, Giao tiếp, Thái độ")
    print("   - Đã được thiết kế để có CR < 0.1 (nhất quán)")
    print()
    print("3. File ma trận ứng viên 3x3 mẫu:")
    print(f"   - {candidate_matrix_file}")
    print("   - Ma trận so sánh 3 ứng viên theo tiêu chí 'Kinh nghiệm'")
    print("   - Có thể dùng làm mẫu cho các tiêu chí khác")
    print()
    print("CẤU TRÚC WORKFLOW:")
    print("1. Tạo đợt tuyển dụng mới")
    print("2. Chọn 5 tiêu chí")
    print("3. Import ma trận tiêu chí từ file test_criteria_matrix_5x5.xlsx")
    print("4. Import danh sách ứng viên từ file test_candidates_3.xlsx")
    print("5. Nhập ma trận so sánh ứng viên cho từng tiêu chí")
    print("   (có thể dùng file test_candidate_matrix_3x3_kinh_nghiem.xlsx làm mẫu)")

if __name__ == "__main__":
    main()
