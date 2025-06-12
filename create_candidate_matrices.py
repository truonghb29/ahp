#!/usr/bin/env python3
"""
Tạo các file ma trận test cho ứng viên theo từng tiêu chí
"""
from openpyxl import Workbook

def create_candidate_matrix_3x3(criterion_name, matrix_data, file_name):
    """Tạo ma trận so sánh ứng viên 3x3 cho một tiêu chí"""
    wb = Workbook()
    ws = wb.active
    ws.title = f"Ma tran ung vien {criterion_name}"
    
    # Tên ứng viên (có thể thay đổi theo ý muốn)
    candidates = ["Ứng viên A", "Ứng viên B", "Ứng viên C"]
    
    # Thêm header với tên ứng viên (hàng 1)
    for j, candidate in enumerate(candidates, start=2):
        ws.cell(row=1, column=j, value=candidate)
    
    # Thêm tên ứng viên ở cột đầu và ma trận
    for i, (candidate, row_data) in enumerate(zip(candidates, matrix_data), start=2):
        ws.cell(row=i, column=1, value=candidate)
        for j, value in enumerate(row_data, start=2):
            ws.cell(row=i, column=j, value=value)
    
    wb.save(file_name)
    print(f"Đã tạo file {file_name}")
    wb.close()

def main():
    """Tạo 4 ma trận test cho các tiêu chí còn lại"""
    
    # Ma trận cho tiêu chí "Thái độ" - ưu tiên cân bằng
    matrix_thai_do = [
        [1, 2, 1/2],
        [1/2, 1, 1/3],
        [2, 3, 1]
    ]
    
    # Ma trận cho tiêu chí "Kỹ năng giao tiếp" - ưu tiên ứng viên C
    matrix_giao_tiep = [
        [1, 1/2, 1/3],
        [2, 1, 1/2],
        [3, 2, 1]
    ]
    
    # Ma trận cho tiêu chí "Khả năng học hỏi" - ưu tiên ứng viên A
    matrix_hoc_hoi = [
        [1, 3, 4],
        [1/3, 1, 2],
        [1/4, 1/2, 1]
    ]
    
    # Ma trận cho tiêu chí "Làm việc nhóm" - ưu tiên ứng viên B
    matrix_nhom = [
        [1, 1/2, 2],
        [2, 1, 3],
        [1/2, 1/3, 1]
    ]
    
    # Tạo các file
    create_candidate_matrix_3x3("thai_do", matrix_thai_do, "test_candidate_matrix_3x3_thai_do.xlsx")
    create_candidate_matrix_3x3("giao_tiep", matrix_giao_tiep, "test_candidate_matrix_3x3_giao_tiep.xlsx")
    create_candidate_matrix_3x3("hoc_hoi", matrix_hoc_hoi, "test_candidate_matrix_3x3_hoc_hoi.xlsx")
    create_candidate_matrix_3x3("nhom", matrix_nhom, "test_candidate_matrix_3x3_nhom.xlsx")
    
    print("\n✅ Đã tạo thành công 4 ma trận test cho ứng viên!")
    print("Các file được tạo:")
    print("- test_candidate_matrix_3x3_thai_do.xlsx")
    print("- test_candidate_matrix_3x3_giao_tiep.xlsx")
    print("- test_candidate_matrix_3x3_hoc_hoi.xlsx")
    print("- test_candidate_matrix_3x3_nhom.xlsx")

if __name__ == "__main__":
    main()
