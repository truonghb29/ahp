from flask import Flask, render_template, request, redirect, url_for, send_file, session, flash, jsonify
from flask_sqlalchemy import SQLAlchemy
import json
import numpy as np
from datetime import datetime
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')  # Sử dụng Agg backend cho matplotlib
import io
import base64
from matplotlib.backends.backend_agg import FigureCanvasAgg as FigureCanvas
import pandas as pd
import seaborn as sns
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference, PieChart
from werkzeug.utils import secure_filename
import os
import tempfile
from fpdf import FPDF

app = Flask(__name__)
app.secret_key = '105008truonganhminhtrong' # Thay bằng chuỗi bí mật thực sự

# Cấu hình upload file
UPLOAD_FOLDER = 'uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Cấu hình kết nối MySQL
# app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql+pymysql://root:Theanh412%40@localhost:3306/recruitment_ahp'
# app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql+pymysql://root:123456@localhost:3306/recruitment_ahp'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# Định nghĩa các model
class RecruitmentRound(db.Model):
    __tablename__ = 'recruitment_rounds'
    round_id = db.Column(db.Integer, primary_key=True)
    round_name = db.Column(db.String(255), nullable=False)
    description = db.Column(db.Text)
    position = db.Column(db.String(100), nullable=False)
    created_at = db.Column(db.DateTime, server_default=db.func.now())

class RecruitmentCriteria(db.Model):
    __tablename__ = 'recruitment_criteria'
    criterion_id = db.Column(db.Integer, primary_key=True)
    round_id = db.Column(db.Integer, db.ForeignKey('recruitment_rounds.round_id'), nullable=False)
    criterion_name = db.Column(db.String(100), nullable=False)
    # is_custom = db.Column(db.Boolean, default=False) # Cột này có thể không còn cần thiết
    created_at = db.Column(db.DateTime, server_default=db.func.now())

class CriteriaMatrix(db.Model):
    __tablename__ = 'criteria_matrix'
    matrix_id = db.Column(db.Integer, primary_key=True)
    round_id = db.Column(db.Integer, db.ForeignKey('recruitment_rounds.round_id'), nullable=False)
    matrix_data = db.Column(db.JSON, nullable=False)
    consistency_ratio = db.Column(db.Float, nullable=False)
    created_at = db.Column(db.DateTime, server_default=db.func.now())

class Candidate(db.Model):
    __tablename__ = 'candidates'
    candidate_id = db.Column(db.Integer, primary_key=True)
    round_id = db.Column(db.Integer, db.ForeignKey('recruitment_rounds.round_id'), nullable=False)
    full_name = db.Column(db.String(255), nullable=False)
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, server_default=db.func.now())

class CandidateScore(db.Model):
    __tablename__ = 'candidate_scores'
    score_id = db.Column(db.Integer, primary_key=True)
    round_id = db.Column(db.Integer, db.ForeignKey('recruitment_rounds.round_id'), nullable=False)
    candidate_id = db.Column(db.Integer, db.ForeignKey('candidates.candidate_id'), nullable=False)
    total_score = db.Column(db.Float, nullable=False)
    ranking = db.Column(db.Integer, nullable=False)
    created_at = db.Column(db.DateTime, server_default=db.func.now())

# --- CÁC HÀM HELPER VÀ GIÁ TRỊ RI ---
ri_values = {
    1: 0.00, 2: 0.00, 3: 0.58, 4: 0.90, 5: 1.12, 6: 1.24,
    7: 1.32, 8: 1.41, 9: 1.45, 10: 1.49, 11: 1.51, 12: 1.48,
    13: 1.56, 14: 1.57, 15: 1.59
}

# --- CÁC HÀM HELPER CHO EXCEL ---
def read_candidates_from_excel(file_path):
    """Đọc danh sách ứng viên từ file Excel"""
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        
        candidates = []
        # Đọc từ cột A, bắt đầu từ hàng 2 (hàng 1 là header)
        for row in range(2, ws.max_row + 1):
            cell_value = ws[f'A{row}'].value
            if cell_value and str(cell_value).strip():
                candidates.append(str(cell_value).strip())
        
        wb.close()
        return candidates, None
    except Exception as e:
        return None, str(e)

def read_matrix_from_excel(file_path, expected_size=None):
    """Đọc ma trận so sánh từ file Excel"""
    wb = None
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        
        matrix = []
        
        # Kiểm tra xem có header không (A1 trống và B1, C1... có giá trị)
        has_header = (ws.cell(row=1, column=1).value is None or 
                     str(ws.cell(row=1, column=1).value).strip() == "")
        
        # Nếu có header, ma trận bắt đầu từ B2, ngược lại từ A1
        start_row = 2 if has_header else 1
        start_col = 2 if has_header else 1
        
        # Tính toán kích thước ma trận
        if expected_size:
            actual_size = expected_size
        else:
            # Tự động phát hiện kích thước ma trận
            if has_header:
                actual_size = min(ws.max_row - 1, ws.max_column - 1)
            else:
                actual_size = min(ws.max_row, ws.max_column)
        
        # Đọc ma trận
        for row in range(start_row, start_row + actual_size):
            matrix_row = []
            for col in range(start_col, start_col + actual_size):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value is None or cell_value == '':
                    matrix_row.append(1.0)  # Giá trị mặc định
                else:
                    try:
                        matrix_row.append(float(cell_value))
                    except (ValueError, TypeError):
                        matrix_row.append(1.0)
            matrix.append(matrix_row)
        
        # Kiểm tra ma trận vuông
        if matrix and len(matrix) != len(matrix[0]):
            return None, "Ma trận không phải ma trận vuông"
        
        # Kiểm tra kích thước nếu có expected_size
        if expected_size and len(matrix) != expected_size:
            return None, f"Ma trận có kích thước {len(matrix)}x{len(matrix[0])}, cần {expected_size}x{expected_size}"
        
        return matrix, None
    except Exception as e:
        return None, str(e)
    finally:
        if wb:
            wb.close()

def read_criteria_names_from_excel(file_path, expected_size=None):
    """Đọc tên tiêu chí từ file Excel (từ hàng 1 và cột 1)"""
    wb = None
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        
        criteria_names = []
        
        # Phương pháp 1: Đọc từ hàng 1, bỏ qua cột 1 (A1) vì thường để trống cho ma trận có header
        # Đọc từ cột 2 trở đi (B1, C1, D1, ...)
        max_col = ws.max_column
        header_names = []
        
        for col in range(2, max_col + 1):  # Bắt đầu từ cột 2 (B)
            cell_value = ws.cell(row=1, column=col).value
            if cell_value and str(cell_value).strip():
                header_names.append(str(cell_value).strip())
        
        # Nếu đọc được đủ tên tiêu chí từ header và phù hợp với expected_size
        if header_names and (not expected_size or len(header_names) == expected_size):
            criteria_names = header_names
        else:
            # Phương pháp 2: Đọc từ cột 1, bỏ qua hàng 1 (A2, A3, A4, ...)
            # Đọc từ hàng 2 trở đi
            max_row = ws.max_row
            col_names = []
            
            for row in range(2, max_row + 1):  # Bắt đầu từ hàng 2
                cell_value = ws.cell(row=row, column=1).value
                if cell_value and str(cell_value).strip():
                    col_names.append(str(cell_value).strip())
            
            # Sử dụng tên từ cột nếu phù hợp
            if col_names and (not expected_size or len(col_names) == expected_size):
                criteria_names = col_names
        
        # Nếu vẫn không đọc được hoặc không đủ số lượng, tạo tên mặc định
        if not criteria_names or (expected_size and len(criteria_names) != expected_size):
            if expected_size:
                criteria_names = [f'Tiêu chí {i+1}' for i in range(expected_size)]
            else:
                criteria_names = [f'Tiêu chí {i+1}' for i in range(6)]  # Mặc định 6 tiêu chí
        
        # Đảm bảo đúng số lượng
        if expected_size and len(criteria_names) > expected_size:
            criteria_names = criteria_names[:expected_size]
        elif expected_size and len(criteria_names) < expected_size:
            # Bổ sung tên mặc định nếu thiếu
            for i in range(len(criteria_names), expected_size):
                criteria_names.append(f'Tiêu chí {i+1}')
        
        return criteria_names, None
    except Exception as e:
        return None, str(e)
    finally:
        if wb:
            wb.close()

def create_candidate_template_excel():
    """Tạo file Excel mẫu cho danh sách ứng viên"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Danh sach ung vien"
    
    # Header
    ws['A1'] = "Tên ứng viên"
    
    # Dữ liệu mẫu
    sample_candidates = [
        "Nguyễn Văn A",
        "Trần Thị B", 
        "Lê Văn C",
        "Phạm Thị D"
    ]
    
    for i, candidate in enumerate(sample_candidates, start=2):
        ws[f'A{i}'] = candidate
    
    return wb

def create_matrix_template_excel(size, matrix_type="criteria"):
    """Tạo file Excel mẫu cho ma trận so sánh"""
    wb = Workbook()
    ws = wb.active
    ws.title = f"Ma tran so sanh {matrix_type}"
    
    # Tạo ma trận đơn vị với giá trị mẫu
    for i in range(1, size + 1):
        for j in range(1, size + 1):
            if i == j:
                ws.cell(row=i, column=j, value=1)
            elif i < j:
                # Giá trị mẫu cho nửa trên của ma trận
                ws.cell(row=i, column=j, value=1)
            else:
                # Giá trị nghịch đảo cho nửa dưới
                upper_val = ws.cell(row=j, column=i).value or 1
                ws.cell(row=i, column=j, value=1/upper_val)
    
    return wb

def calculate_priority_vector(matrix):
    matrix = np.array(matrix, dtype=float) # Đảm bảo kiểu float
    if matrix.ndim != 2 or matrix.shape[0] != matrix.shape[1] or matrix.shape[0] == 0:
        return np.array([])
    
    col_sums = np.sum(matrix, axis=0)
    if np.any(col_sums == 0): # Tránh chia cho 0
        return np.full(matrix.shape[0], 1/matrix.shape[0]) # Default to equal weights as a fallback

    normalized_matrix = matrix / col_sums
    weights = np.mean(normalized_matrix, axis=1)
    return weights

def calculate_consistency_ratio(matrix):
    matrix = np.array(matrix, dtype=float) # Đảm bảo kiểu float
    n = matrix.shape[0]

    if n == 0:
        return 0.0 
    if n <= 2: # Ma trận 1x1 hoặc 2x2 luôn nhất quán
        return 0.0

    weights = calculate_priority_vector(matrix)
    if weights.size == 0 or weights.shape[0] != n: # Kiểm tra weights hợp lệ
        return float('inf') 

    aw = np.dot(matrix, weights)
    
    if np.any(weights == 0):
        lambda_max = n 
    else:
        lambda_max = np.mean(aw / weights)

    if n - 1 == 0: 
        return 0.0 if lambda_max == n else float('inf')
        
    ci = (lambda_max - n) / (n - 1)
    
    ri = ri_values.get(n)
    if ri is None: 
        ri = ri_values.get(15, 1.59) # Mặc định cho n > 15
    
    if ri == 0: # Chỉ xảy ra nếu n <= 2 và logic ở trên chưa bắt được, hoặc ri_values bị sửa đổi
        return 0.0 if ci == 0 else float('inf') # Nếu CI cũng là 0 thì nhất quán, ngược lại là lỗi
             
    cr = ci / ri
    return cr

def calculate_detailed_consistency_metrics(matrix):
    """Tính toán chi tiết các chỉ số nhất quán: Lambda max, CI, CR và vector nhất quán"""
    matrix = np.array(matrix, dtype=float)
    n = matrix.shape[0]
    
    if n == 0:
        return {
            'lambda_max': 0.0,
            'ci': 0.0,
            'cr': 0.0,
            'consistency_vector': np.array([]),
            'ri': 0.0
        }
    
    if n <= 2:
        weights = calculate_priority_vector(matrix)
        return {
            'lambda_max': n,
            'ci': 0.0,
            'cr': 0.0,
            'consistency_vector': weights,
            'ri': 0.0
        }
    
    weights = calculate_priority_vector(matrix)
    if weights.size == 0 or weights.shape[0] != n:
        return {
            'lambda_max': float('inf'),
            'ci': float('inf'),
            'cr': float('inf'),
            'consistency_vector': np.array([]),
            'ri': ri_values.get(n, 1.59)
        }
    
    # Tính lambda_max
    aw = np.dot(matrix, weights)
    if np.any(weights == 0):
        lambda_max = n
    else:
        lambda_max = np.mean(aw / weights)
    
    # Tính CI (Consistency Index)
    if n - 1 == 0:
        ci = 0.0
    else:
        ci = (lambda_max - n) / (n - 1)
    
    # Lấy RI (Random Index)
    ri = ri_values.get(n)
    if ri is None:
        ri = ri_values.get(15, 1.59)
    
    # Tính CR (Consistency Ratio)
    if ri == 0:
        cr = 0.0 if ci == 0 else float('inf')
    else:
        cr = ci / ri
    
    return {
        'lambda_max': lambda_max,
        'ci': ci,
        'cr': cr,
        'consistency_vector': weights,
        'ri': ri
    }

# --- CÁC BIẾN TOÀN CỤC CŨ (Đã comment out hoặc sẽ được thay thế) ---
# criteria = ["Kiến thức chuyên môn", ...] 
# num_criteria_global = len(criteria) # (Sử dụng tên khác để tránh nhầm lẫn với num_criteria trong session)
# default_pairwise = np.array([...])
# optimal_values = np.array([100, 10, 10, 10, 5, 10]) # Biến này có thể không còn dùng trong quy trình AHP mới

# Kiểm tra kết nối đến database
@app.route('/check_connection')
def check_connection():
    try:
        RecruitmentRound.query.first()
        return "Database connection is successful."
    except Exception as e:
        return f"Database connection failed: {str(e)}"

# Trang chính: Tạo đợt tuyển dụng
@app.route('/', methods=['GET', 'POST'])
def index_or_create_round():
    if request.method == 'POST':
        round_name = request.form.get('round_name','').strip()
        description = request.form.get('description', '').strip()
        position = request.form.get('position','').strip()

        if not round_name or not position:
            flash("Tên đợt tuyển dụng và vị trí không được để trống.", "danger")
            return render_template('index.html', round_name=round_name, description=description, position=position)

        try:
            new_round = RecruitmentRound(
                round_name=round_name,
                description=description,
                position=position
            )
            db.session.add(new_round)
            db.session.commit()
            
            session.clear() # Xóa session cũ khi bắt đầu đợt mới
            session['round_id'] = new_round.round_id
            session['round_name'] = new_round.round_name
            flash(f"Đợt tuyển dụng '{round_name}' đã được tạo thành công!", "success")
            return redirect(url_for('setup_criteria_count'))
        except Exception as e:
            db.session.rollback()
            flash(f"Lỗi khi tạo đợt tuyển dụng: {str(e)}", "danger")
            return render_template('index.html', round_name=round_name, description=description, position=position)
    
    # Xóa session khi truy cập trang chủ bằng GET để bắt đầu mới (tuỳ chọn)
    # session.pop('round_id', None) 
    # session.pop('round_name', None)
    # ... xóa các key khác của AHP process
    return render_template('index.html')

@app.route('/setup_criteria_count', methods=['GET', 'POST'])
def setup_criteria_count():
    if 'round_id' not in session:
        flash("Vui lòng tạo hoặc chọn một đợt tuyển dụng trước.", "warning")
        return redirect(url_for('index_or_create_round'))

    if request.method == 'POST':
        try:
            num_criteria = int(request.form['num_criteria'])
            if num_criteria <= 1:
                flash("Số lượng tiêu chí phải lớn hơn 1.", "warning")
                return render_template('setup_criteria_count.html', 
                                       round_name=session.get('round_name'),
                                       num_criteria_value=request.form.get('num_criteria'))
            session['num_criteria'] = num_criteria
            return redirect(url_for('setup_criteria_details'))
        except ValueError:
            flash("Vui lòng nhập một số hợp lệ cho số lượng tiêu chí.", "danger")
            return render_template('setup_criteria_count.html', 
                                   round_name=session.get('round_name'))
                                   
    return render_template('setup_criteria_count.html', round_name=session.get('round_name'))

@app.route('/setup_criteria_details', methods=['GET', 'POST'])
def setup_criteria_details():
    if 'round_id' not in session or 'num_criteria' not in session:
        flash("Thông tin đợt tuyển dụng hoặc số lượng tiêu chí bị thiếu. Vui lòng thử lại.", "warning")
        return redirect(url_for('index_or_create_round'))

    num_criteria = session.get('num_criteria', 0)
    if num_criteria == 0: # Đảm bảo num_criteria hợp lệ
        return redirect(url_for('setup_criteria_count'))

    # Lấy giá trị đã nhập nếu có (khi render lại do lỗi)
    criteria_names_input = [''] * num_criteria
    matrix_values_input = np.ones((num_criteria, num_criteria)).tolist()


    if request.method == 'POST':
        criteria_names = [request.form.get(f'criterion_name_{i}', '').strip() for i in range(num_criteria)]
        if any(not name for name in criteria_names):
            flash("Tên tiêu chí không được để trống.", "danger")
            # Truyền lại giá trị đã nhập để người dùng không phải nhập lại tất cả
            for i in range(num_criteria): matrix_values_input[i] = [request.form.get(f'criteria_pairwise_{i}_{j}', 1.0) for j in range(num_criteria)]
            return render_template('setup_criteria_details.html', 
                                   num_criteria=num_criteria, 
                                   round_name=session.get('round_name'),
                                   criteria_names_input=criteria_names, # tên đã nhập
                                   matrix_values_input=matrix_values_input # ma trận đã nhập
                                   )

        pairwise_matrix_criteria_list = []
        try:
            for i in range(num_criteria):
                row = []
                for j in range(num_criteria):
                    value_str = request.form.get(f'criteria_pairwise_{i}_{j}', '1.0') # Mặc định là 1 nếu thiếu
                    val = float(value_str)
                    if val <= 0: # Giá trị phải dương
                        flash(f"Giá trị so sánh giữa '{criteria_names[i]}' và '{criteria_names[j]}' phải là số dương.", "danger")
                        for r_idx in range(num_criteria): matrix_values_input[r_idx] = [request.form.get(f'criteria_pairwise_{r_idx}_{c_idx}', 1.0) for c_idx in range(num_criteria)]
                        return render_template('setup_criteria_details.html', num_criteria=num_criteria, criteria_names_input=criteria_names, matrix_values_input=matrix_values_input, round_name=session.get('round_name'))
                    row.append(val)
                pairwise_matrix_criteria_list.append(row)
        except ValueError:
            flash("Giá trị trong ma trận không hợp lệ. Vui lòng nhập số.", "danger")
            # Truyền lại giá trị đã nhập
            for i in range(num_criteria): matrix_values_input[i] = [request.form.get(f'criteria_pairwise_{i}_{j}', 1.0) for j in range(num_criteria)]
            return render_template('setup_criteria_details.html', num_criteria=num_criteria, criteria_names_input=criteria_names, matrix_values_input=matrix_values_input, round_name=session.get('round_name'))

        pairwise_matrix_criteria = np.array(pairwise_matrix_criteria_list)
        
        cr_criteria = calculate_consistency_ratio(pairwise_matrix_criteria)
        weights_criteria = calculate_priority_vector(pairwise_matrix_criteria)

        session['criteria_names'] = criteria_names
        session['pairwise_matrix_criteria'] = pairwise_matrix_criteria.tolist()
        session['weights_criteria'] = weights_criteria.tolist()
        session['cr_criteria'] = cr_criteria
        
        round_id = session['round_id']
        # Xóa tiêu chí cũ của round này trước khi thêm mới (nếu user quay lại và sửa)
        RecruitmentCriteria.query.filter_by(round_id=round_id).delete()
        for name in criteria_names:
            criterion_db = RecruitmentCriteria(round_id=round_id, criterion_name=name)
            db.session.add(criterion_db)
        
        existing_matrix = CriteriaMatrix.query.filter_by(round_id=round_id).first()
        matrix_data_json = json.dumps({
            "matrix": pairwise_matrix_criteria.tolist(),
            "weights": weights_criteria.tolist(),
            "criteria_names_at_creation": criteria_names
        })
        if existing_matrix:
            existing_matrix.matrix_data = matrix_data_json
            existing_matrix.consistency_ratio = cr_criteria
        else:
            new_criteria_matrix_db = CriteriaMatrix(
                round_id=round_id,
                matrix_data=matrix_data_json,
                consistency_ratio=cr_criteria
            )
            db.session.add(new_criteria_matrix_db)
        db.session.commit()

        if cr_criteria < 0.1:
            flash(f"Ma trận tiêu chí nhất quán (CR = {cr_criteria:.4f}).", "success")
            return redirect(url_for('setup_candidates_count'))
        else:
            flash(f"CR của ma trận tiêu chí ({cr_criteria:.4f}) >= 0.1. Vui lòng kiểm tra lại.", "warning")
            return render_template('setup_criteria_details.html', 
                                   num_criteria=num_criteria, 
                                   criteria_names_input=criteria_names, 
                                   matrix_values_input=pairwise_matrix_criteria.tolist(), 
                                   round_name=session.get('round_name'))
      # Cho GET request
    # Nếu có giá trị cũ trong session (ví dụ user back hoặc sau khi import), hiển thị lại
    if 'criteria_names' in session and len(session['criteria_names']) == num_criteria:
        criteria_names_input = session['criteria_names']
        print(f"DEBUG: Loaded criteria_names from session: {criteria_names_input}")  # Debug log
    else:
        print(f"DEBUG: No criteria_names in session or wrong length. Session has: {session.get('criteria_names', 'None')}")  # Debug log
        
    if 'pairwise_matrix_criteria' in session and len(session['pairwise_matrix_criteria']) == num_criteria:
        matrix_values_input = session['pairwise_matrix_criteria']
        print(f"DEBUG: Loaded matrix from session with size: {len(matrix_values_input)}x{len(matrix_values_input[0]) if matrix_values_input else 0}")  # Debug log

    return render_template('setup_criteria_details.html', 
                           num_criteria=num_criteria, 
                           criteria_names_input=criteria_names_input,
                           matrix_values_input=matrix_values_input,
                           round_name=session.get('round_name'))

@app.route('/setup_candidates_count', methods=['GET', 'POST'])
def setup_candidates_count():
    if 'round_id' not in session:
        return redirect(url_for('index_or_create_round'))
    if session.get('cr_criteria', 1.0) >= 0.1:
        flash("Ma trận tiêu chí chưa nhất quán. Vui lòng sửa trước khi tiếp tục.", "warning")
        return redirect(url_for('setup_criteria_details'))

    if request.method == 'POST':
        try:
            num_candidates = int(request.form['num_candidates'])
            if num_candidates <= 1:
                flash("Số lượng ứng viên phải lớn hơn 1.", "warning")
                return render_template('setup_candidates_count.html', 
                                       round_name=session.get('round_name'),
                                       criteria_names=session.get('criteria_names'),
                                       num_candidates_value = request.form.get('num_candidates'))
            session['num_candidates'] = num_candidates
            return redirect(url_for('setup_candidate_names'))
        except ValueError:
            flash("Vui lòng nhập một số hợp lệ cho số lượng ứng viên.", "danger")
            return render_template('setup_candidates_count.html', 
                                   round_name=session.get('round_name'),
                                   criteria_names=session.get('criteria_names'))

    return render_template('setup_candidates_count.html', 
                           round_name=session.get('round_name'),
                           criteria_names=session.get('criteria_names'))

@app.route('/setup_candidate_names', methods=['GET', 'POST'])
def setup_candidate_names():
    if 'round_id' not in session or 'num_candidates' not in session:
        return redirect(url_for('setup_candidates_count'))
    
    num_candidates = session.get('num_candidates', 0)
    if num_candidates == 0:
        return redirect(url_for('setup_candidates_count'))

    candidate_names_input = [''] * num_candidates

    if request.method == 'POST':
        candidate_names = [request.form.get(f'candidate_name_{i}', '').strip() for i in range(num_candidates)]
        if any(not name for name in candidate_names):
            flash("Tên ứng viên không được để trống.", "danger")
            return render_template('setup_candidate_names.html', 
                                   num_candidates=num_candidates, 
                                   round_name=session.get('round_name'),
                                   candidate_names_input=candidate_names)

        session['candidate_names'] = candidate_names
        session['candidate_pairwise_matrices_details'] = [None] * len(session.get('criteria_names', []))
        
        round_id = session['round_id']
        # Xóa ứng viên cũ của round này trước khi thêm mới
        Candidate.query.filter_by(round_id=round_id).delete() # Cẩn thận với việc xóa dữ liệu nếu có liên kết
        db.session.commit() # Commit việc xóa trước khi thêm

        candidate_ids = []
        for name in candidate_names:
            candidate_db = Candidate(round_id=round_id, full_name=name)
            db.session.add(candidate_db)
            db.session.flush() 
            candidate_ids.append(candidate_db.candidate_id)
        db.session.commit()
        session['candidate_ids'] = candidate_ids

        return redirect(url_for('input_candidate_comparison_for_criterion', criterion_idx=0))
    
    # Cho GET request
    if 'candidate_names' in session and len(session['candidate_names']) == num_candidates:
        candidate_names_input = session['candidate_names']
       
    return render_template('setup_candidate_names.html', 
                           num_candidates=num_candidates, 
                           round_name=session.get('round_name'),
                           candidate_names_input=candidate_names_input)

@app.route('/import_candidates', methods=['GET', 'POST'])
def import_candidates():
    if 'round_id' not in session:
        return redirect(url_for('index_or_create_round'))
    
    if request.method == 'POST':
        # Kiểm tra xem có file được upload không
        if 'file' not in request.files:
            flash('Không có file được chọn', 'danger')
            return render_template('import_candidates.html', round_name=session.get('round_name'))
        
        file = request.files['file']
        if file.filename == '':
            flash('Không có file được chọn', 'danger')
            return render_template('import_candidates.html', round_name=session.get('round_name'))
        
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(file_path)
            
            try:
                # Đọc danh sách ứng viên từ Excel
                candidates, error = read_candidates_from_excel(file_path)
                
                if error:
                    flash(f'Lỗi đọc file Excel: {error}', 'danger')
                    return render_template('import_candidates.html', round_name=session.get('round_name'))
                
                if not candidates:
                    flash('File Excel không chứa ứng viên nào hoặc định dạng không đúng', 'warning')
                    return render_template('import_candidates.html', round_name=session.get('round_name'))
                
                if len(candidates) < 2:
                    flash('Cần ít nhất 2 ứng viên để tiến hành so sánh', 'warning')
                    return render_template('import_candidates.html', round_name=session.get('round_name'))
                
                # Lưu vào session
                session['num_candidates'] = len(candidates)
                session['candidate_names'] = candidates
                session['candidate_pairwise_matrices_details'] = [None] * len(session.get('criteria_names', []))
                
                # Lưu vào database
                round_id = session['round_id']
                Candidate.query.filter_by(round_id=round_id).delete()
                db.session.commit()
                
                candidate_ids = []
                for name in candidates:
                    candidate_db = Candidate(round_id=round_id, full_name=name)
                    db.session.add(candidate_db)
                    db.session.flush()
                    candidate_ids.append(candidate_db.candidate_id)
                db.session.commit()
                session['candidate_ids'] = candidate_ids
                
                flash(f'Đã import thành công {len(candidates)} ứng viên từ file Excel!', 'success')
                return redirect(url_for('input_candidate_comparison_for_criterion', criterion_idx=0))
                
            except Exception as e:
                flash(f'Lỗi xử lý file: {str(e)}', 'danger')
                return render_template('import_candidates.html', round_name=session.get('round_name'))
            finally:
                # Xóa file tạm
                if os.path.exists(file_path):
                    os.remove(file_path)
        else:
            flash('File không hợp lệ. Chỉ chấp nhận file .xlsx hoặc .xls', 'danger')
    
    return render_template('import_candidates.html', round_name=session.get('round_name'))

@app.route('/download_candidate_template')
def download_candidate_template():
    """Download file Excel mẫu cho danh sách ứng viên"""
    try:
        wb = create_candidate_template_excel()
        
        # Tạo file tạm
        temp_path = os.path.join(app.config['UPLOAD_FOLDER'], 'mau_danh_sach_ung_vien.xlsx')
        wb.save(temp_path)
        wb.close()
        
        return send_file(temp_path, 
                        as_attachment=True, 
                        download_name='mau_danh_sach_ung_vien.xlsx',
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        flash(f'Lỗi tạo file mẫu: {str(e)}', 'danger')
        return redirect(url_for('setup_candidates_count'))
    finally:
        # Xóa file tạm sau khi gửi
        if 'temp_path' in locals() and os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except:
                pass


@app.route('/input_candidate_comparison_for_criterion/<int:criterion_idx>', methods=['GET', 'POST'])
def input_candidate_comparison_for_criterion(criterion_idx):
    required_sessions = ['round_id', 'criteria_names', 'candidate_names', 'candidate_pairwise_matrices_details']
    for key in required_sessions:
        if key not in session:
            flash(f"Thiếu thông tin cần thiết ({key}). Vui lòng bắt đầu lại.", "warning")
            return redirect(url_for('index_or_create_round'))

    criteria_names = session.get('criteria_names', [])
    candidate_names = session.get('candidate_names', [])
    num_candidates = len(candidate_names)
    candidate_matrices_details = session.get('candidate_pairwise_matrices_details', [])

    # Đảm bảo candidate_matrices_details có đủ độ dài
    if len(candidate_matrices_details) != len(criteria_names):
        candidate_matrices_details = [None] * len(criteria_names)
        session['candidate_pairwise_matrices_details'] = candidate_matrices_details


    if not (0 <= criterion_idx < len(criteria_names)):
        # Kiểm tra nếu đã hoàn thành tất cả các tiêu chí
        all_done_and_consistent = True
        if len(candidate_matrices_details) != len(criteria_names): # Chưa đủ số ma trận
            all_done_and_consistent = False
        else:
            for i, detail in enumerate(candidate_matrices_details):
                if detail is None: # Một tiêu chí chưa có ma trận
                     # flash(f"Vui lòng hoàn thành ma trận so sánh cho tiêu chí '{criteria_names[i]}'.", "info")
                     return redirect(url_for('input_candidate_comparison_for_criterion', criterion_idx=i))
                if detail['cr'] >= 0.1:
                    all_done_and_consistent = False # Vẫn cho qua nhưng sẽ báo lỗi ở trang kết quả

        if all_done_and_consistent or (all(d is not None for d in candidate_matrices_details)): # Nếu đã điền hết (dù có CR lỗi)
            return redirect(url_for('calculate_final_ranking'))
        else: # Trường hợp lạ, redirect về trang đầu
            flash("Có lỗi trong quy trình, vui lòng bắt đầu lại.", "danger")
            return redirect(url_for('index_or_create_round'))


    current_criterion_name = criteria_names[criterion_idx]
    matrix_values_input = np.ones((num_candidates, num_candidates)).tolist() # Default

    if request.method == 'POST':
        pairwise_matrix_candidate_list = []
        try:
            for i in range(num_candidates):
                row = []
                for j in range(num_candidates):
                    value_str = request.form.get(f'candidate_pairwise_{i}_{j}', '1.0')
                    val = float(value_str)
                    if val <= 0:
                        flash(f"Giá trị so sánh giữa '{candidate_names[i]}' và '{candidate_names[j]}' phải là số dương.", "danger")
                        # Truyền lại giá trị đã nhập
                        for r_idx in range(num_candidates): 
                            matrix_values_input[r_idx] = [request.form.get(f'candidate_pairwise_{r_idx}_{c_idx}', 1.0) for c_idx in range(num_candidates)]
                        return render_template('input_candidate_comparison.html', criterion_name=current_criterion_name, criterion_idx=criterion_idx, candidate_names=candidate_names, matrix_values_input=matrix_values_input, round_name=session.get('round_name'))
                    row.append(val)
                pairwise_matrix_candidate_list.append(row)
        except ValueError:
            flash("Giá trị trong ma trận không hợp lệ. Vui lòng nhập số.", "danger")
            # Truyền lại giá trị đã nhập
            for i in range(num_candidates): matrix_values_input[i] = [request.form.get(f'candidate_pairwise_{i}_{j}', 1.0) for j in range(num_candidates)]
            return render_template('input_candidate_comparison.html', criterion_name=current_criterion_name, criterion_idx=criterion_idx, candidate_names=candidate_names, matrix_values_input=matrix_values_input, round_name=session.get('round_name'))

        pairwise_matrix_candidate = np.array(pairwise_matrix_candidate_list)
        cr_candidate = calculate_consistency_ratio(pairwise_matrix_candidate)
        weights_candidate_local = calculate_priority_vector(pairwise_matrix_candidate)

        candidate_matrices_details[criterion_idx] = {
            'matrix': pairwise_matrix_candidate.tolist(),
            'cr': cr_candidate,
            'weights': weights_candidate_local.tolist(),
            'criterion_name': current_criterion_name
        }
        session['candidate_pairwise_matrices_details'] = candidate_matrices_details        
        if cr_candidate >= 0.1:
            flash(f"CR cho ứng viên theo tiêu chí '{current_criterion_name}' ({cr_candidate:.4f}) >= 0.1. Vui lòng kiểm tra lại.", "warning")            
            return render_template('input_candidate_comparison.html',
                                   criterion_name=current_criterion_name,
                                   criterion_idx=criterion_idx,
                                   candidate_names=candidate_names,
                                   matrix_values_input=pairwise_matrix_candidate.tolist(),
                                   round_name=session.get('round_name'))
        
        flash(f"Ma trận cho tiêu chí '{current_criterion_name}' đã được lưu (CR = {cr_candidate:.4f}).", "success")
        next_criterion_idx = criterion_idx + 1
        if next_criterion_idx < len(criteria_names):
            return redirect(url_for('input_candidate_comparison_for_criterion', criterion_idx=next_criterion_idx))
        else:
            return redirect(url_for('calculate_final_ranking'))

    # Cho GET request: lấy lại giá trị đã nhập nếu có
    if candidate_matrices_details[criterion_idx] is not None:
        matrix_values_input = candidate_matrices_details[criterion_idx]['matrix']

    return render_template('input_candidate_comparison.html', 
                           criterion_name=current_criterion_name,
                           criterion_idx=criterion_idx,
                           candidate_names=candidate_names,
                           matrix_values_input=matrix_values_input,
                           round_name=session.get('round_name'))

@app.route('/import_criteria_matrix', methods=['GET', 'POST'])
def import_criteria_matrix():
    if 'round_id' not in session or 'num_criteria' not in session:
        return redirect(url_for('index_or_create_round'))
    
    num_criteria = session.get('num_criteria', 0)
    
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('Không có file được chọn', 'danger')
            return render_template('import_criteria_matrix.html', 
                                 round_name=session.get('round_name'),
                                 num_criteria=num_criteria)
        
        file = request.files['file']
        if file.filename == '':
            flash('Không có file được chọn', 'danger')
            return render_template('import_criteria_matrix.html', 
                                 round_name=session.get('round_name'),
                                 num_criteria=num_criteria)
        
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(file_path)
            try:
                # Đọc ma trận từ Excel
                matrix_data, error = read_matrix_from_excel(file_path, num_criteria)
                
                if error:
                    flash(f'Lỗi đọc file Excel: {error}', 'danger')
                    return render_template('import_criteria_matrix.html', 
                                         round_name=session.get('round_name'),
                                         num_criteria=num_criteria)
                
                if not matrix_data or len(matrix_data) != num_criteria:
                    flash(f'Ma trận không đúng kích thước. Cần ma trận {num_criteria}x{num_criteria}', 'warning')
                    return render_template('import_criteria_matrix.html', 
                                         round_name=session.get('round_name'),
                                         num_criteria=num_criteria)
                
                # Kiểm tra ma trận hợp lệ
                pairwise_matrix_criteria = np.array(matrix_data)
                
                # Kiểm tra các giá trị dương
                if np.any(pairwise_matrix_criteria <= 0):
                    flash('Ma trận chứa giá trị không hợp lệ (phải là số dương)', 'danger')
                    return render_template('import_criteria_matrix.html', 
                                         round_name=session.get('round_name'),
                                         num_criteria=num_criteria)
                
                cr_criteria = calculate_consistency_ratio(pairwise_matrix_criteria)
                weights_criteria = calculate_priority_vector(pairwise_matrix_criteria)
                
                # Đọc tên tiêu chí từ file Excel
                criteria_names_from_excel, criteria_error = read_criteria_names_from_excel(file_path, num_criteria)
                
                if criteria_names_from_excel and not criteria_error:
                    criteria_names = criteria_names_from_excel
                    flash(f'Đã tự động điền tên tiêu chí từ file Excel: {", ".join(criteria_names)}', 'info')
                else:
                    # Fallback về tên mặc định nếu không đọc được
                    criteria_names = session.get('criteria_names')
                    if not criteria_names or len(criteria_names) != num_criteria:
                        criteria_names = [f'Tiêu chí {i+1}' for i in range(num_criteria)]
                    flash('Không thể đọc tên tiêu chí từ file, sử dụng tên mặc định', 'warning')
                
                session['criteria_names'] = criteria_names
                session['pairwise_matrix_criteria'] = pairwise_matrix_criteria.tolist()
                session['weights_criteria'] = weights_criteria.tolist()
                session['cr_criteria'] = cr_criteria
                
                # Lưu vào database
                round_id = session['round_id']
                RecruitmentCriteria.query.filter_by(round_id=round_id).delete()
                for name in criteria_names:
                    criterion_db = RecruitmentCriteria(round_id=round_id, criterion_name=name)
                    db.session.add(criterion_db)
                
                existing_matrix = CriteriaMatrix.query.filter_by(round_id=round_id).first()
                matrix_data_json = json.dumps({
                    "matrix": pairwise_matrix_criteria.tolist(),
                    "weights": weights_criteria.tolist(),
                    "criteria_names_at_creation": criteria_names
                })
                if existing_matrix:
                    existing_matrix.matrix_data = matrix_data_json
                    existing_matrix.consistency_ratio = cr_criteria
                else:
                    new_criteria_matrix_db = CriteriaMatrix(
                        round_id=round_id,
                        matrix_data=matrix_data_json,
                        consistency_ratio=cr_criteria
                    )
                    db.session.add(new_criteria_matrix_db)
                db.session.commit()
                
                if cr_criteria < 0.1:
                    flash(f'Đã import thành công ma trận tiêu chí (CR = {cr_criteria:.4f})', 'success')
                    return redirect(url_for('setup_criteria_details'))
                else:
                    flash(f'Ma trận đã được import nhưng CR ({cr_criteria:.4f}) >= 0.1. Vui lòng kiểm tra lại.', 'warning')
                    return redirect(url_for('setup_criteria_details'))
                
            except Exception as e:
                flash(f'Lỗi xử lý file: {str(e)}', 'danger')
            finally:
                if os.path.exists(file_path):
                    os.remove(file_path)
        else:
            flash('File không hợp lệ. Chỉ chấp nhận file .xlsx hoặc .xls', 'danger')
    
    return render_template('import_criteria_matrix.html', 
                         round_name=session.get('round_name'),
                         num_criteria=num_criteria)

@app.route('/import_candidate_matrix/<int:criterion_idx>', methods=['GET', 'POST'])
def import_candidate_matrix(criterion_idx):
    required_sessions = ['round_id', 'criteria_names', 'candidate_names', 'candidate_pairwise_matrices_details']
    for key in required_sessions:
        if key not in session:
            flash(f"Thiếu thông tin cần thiết ({key}). Vui lòng bắt đầu lại.", "warning")
            return redirect(url_for('index_or_create_round'))
    
    criteria_names = session.get('criteria_names', [])
    candidate_names = session.get('candidate_names', [])
    num_candidates = len(candidate_names)
    
    if not (0 <= criterion_idx < len(criteria_names)):
        flash("Chỉ số tiêu chí không hợp lệ", "danger")
        return redirect(url_for('input_candidate_comparison_for_criterion', criterion_idx=0))
    
    current_criterion_name = criteria_names[criterion_idx]
    
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('Không có file được chọn', 'danger')
            return render_template('import_candidate_matrix.html', 
                                 criterion_name=current_criterion_name,
                                 criterion_idx=criterion_idx,
                                 candidate_names=candidate_names,
                                 round_name=session.get('round_name'))
        
        file = request.files['file']
        if file.filename == '':
            flash('Không có file được chọn', 'danger')
            return render_template('import_candidate_matrix.html', 
                                 criterion_name=current_criterion_name,
                                 criterion_idx=criterion_idx,
                                 candidate_names=candidate_names,
                                 round_name=session.get('round_name'))
        
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(file_path)
            
            try:
                matrix_data, error = read_matrix_from_excel(file_path, num_candidates)
                
                if error:
                    flash(f'Lỗi đọc file Excel: {error}', 'danger')
                    return render_template('import_candidate_matrix.html', 
                                         criterion_name=current_criterion_name,
                                         criterion_idx=criterion_idx,
                                         candidate_names=candidate_names,
                                         round_name=session.get('round_name'))
                
                if not matrix_data or len(matrix_data) != num_candidates:
                    flash(f'Ma trận không đúng kích thước. Cần ma trận {num_candidates}x{num_candidates}', 'warning')
                    return render_template('import_candidate_matrix.html', 
                                         criterion_name=current_criterion_name,
                                         criterion_idx=criterion_idx,
                                         candidate_names=candidate_names,
                                         round_name=session.get('round_name'))
                
                pairwise_matrix_candidate = np.array(matrix_data)
                
                if np.any(pairwise_matrix_candidate <= 0):
                    flash('Ma trận chứa giá trị không hợp lệ (phải là số dương)', 'danger')
                    return render_template('import_candidate_matrix.html', 
                                         criterion_name=current_criterion_name,
                                         criterion_idx=criterion_idx,
                                         candidate_names=candidate_names,
                                         round_name=session.get('round_name'))
                
                cr_candidate = calculate_consistency_ratio(pairwise_matrix_candidate)
                weights_candidate_local = calculate_priority_vector(pairwise_matrix_candidate)
                
                candidate_matrices_details = session.get('candidate_pairwise_matrices_details', [])
                candidate_matrices_details[criterion_idx] = {
                    'matrix': pairwise_matrix_candidate.tolist(),
                    'cr': cr_candidate,
                    'weights': weights_candidate_local.tolist(),
                    'criterion_name': current_criterion_name
                }
                session['candidate_pairwise_matrices_details'] = candidate_matrices_details
                if cr_candidate >= 0.1:
                    flash(f'Ma trận đã được import nhưng CR ({cr_candidate:.4f}) >= 0.1. Vui lòng kiểm tra lại.', 'warning')
                else:
                    flash(f'Đã import thành công ma trận cho tiêu chí "{current_criterion_name}" (CR = {cr_candidate:.4f})', 'success')
                
                # Ở lại trang hiện tại để người dùng có thể xem kết quả tính toán
                return redirect(url_for('input_candidate_comparison_for_criterion', criterion_idx=criterion_idx))
                        
            except Exception as e:
                flash(f'Lỗi xử lý file: {str(e)}', 'danger')
            finally:
                if os.path.exists(file_path):
                    os.remove(file_path)
        else:
            flash('File không hợp lệ. Chỉ chấp nhận file .xlsx hoặc .xls', 'danger')
    
    return render_template('import_candidate_matrix.html', 
                         criterion_name=current_criterion_name,
                         criterion_idx=criterion_idx,
                         candidate_names=candidate_names,
                         round_name=session.get('round_name'))

@app.route('/download_criteria_matrix_template')
def download_criteria_matrix_template():
    """Download file Excel mẫu cho ma trận so sánh tiêu chí"""
    if 'num_criteria' not in session:
        flash('Thông tin số lượng tiêu chí bị thiếu', 'danger')
        return redirect(url_for('setup_criteria_count'))
    
    num_criteria = session.get('num_criteria', 3)
    
    try:
        wb = create_matrix_template_excel(num_criteria, "criteria")
        
        temp_path = os.path.join(app.config['UPLOAD_FOLDER'], f'mau_ma_tran_tieu_chi_{num_criteria}x{num_criteria}.xlsx')
        wb.save(temp_path)
        wb.close()
        
        return send_file(temp_path, 
                        as_attachment=True, 
                        download_name=f'mau_ma_tran_tieu_chi_{num_criteria}x{num_criteria}.xlsx',
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        flash(f'Lỗi tạo file mẫu: {str(e)}', 'danger')
        return redirect(url_for('setup_criteria_details'))
    finally:
        if 'temp_path' in locals() and os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except:
                pass

@app.route('/download_candidate_matrix_template/<int:criterion_idx>')
def download_candidate_matrix_template(criterion_idx):
    """Download file Excel mẫu cho ma trận so sánh ứng viên"""
    if 'candidate_names' not in session:
        flash('Thông tin ứng viên bị thiếu', 'danger')
        return redirect(url_for('setup_candidates_count'))
    
    candidate_names = session.get('candidate_names', [])
    num_candidates = len(candidate_names)
    criteria_names = session.get('criteria_names', [])
    
    if not (0 <= criterion_idx < len(criteria_names)):
        flash('Chỉ số tiêu chí không hợp lệ', 'danger')
        return redirect(url_for('input_candidate_comparison_for_criterion', criterion_idx=0))
    
    criterion_name = criteria_names[criterion_idx]
    
    try:
        wb = create_matrix_template_excel(num_candidates, f"ung_vien_{criterion_name}")
        
        # Thêm tên ứng viên làm header
        ws = wb.active
        # Thêm tên ứng viên ở hàng đầu và cột đầu
        for i, name in enumerate(candidate_names, start=1):
            ws.cell(row=1, column=i+1, value=name)  # Header hàng
            ws.cell(row=i+1, column=1, value=name)  # Header cột
        
        temp_path = os.path.join(app.config['UPLOAD_FOLDER'], f'mau_ma_tran_ung_vien_{criterion_name}_{num_candidates}x{num_candidates}.xlsx')
        wb.save(temp_path)
        wb.close()
        
        return send_file(temp_path, 
                        as_attachment=True, 
                        download_name=f'mau_ma_tran_ung_vien_{criterion_name}_{num_candidates}x{num_candidates}.xlsx',
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        flash(f'Lỗi tạo file mẫu: {str(e)}', 'danger')
        return redirect(url_for('input_candidate_comparison_for_criterion', criterion_idx=criterion_idx))
    finally:
        if 'temp_path' in locals() and os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except:
                pass

@app.route('/calculate_final_ranking')
def calculate_final_ranking():
    required_sessions = ['round_id', 'criteria_names', 'weights_criteria', 'cr_criteria', 
                         'candidate_names', 'candidate_ids', 'candidate_pairwise_matrices_details']
    for key in required_sessions:
        if key not in session:
            flash(f"Thiếu dữ liệu phiên làm việc: {key}. Vui lòng bắt đầu lại.", "danger")
            return redirect(url_for('index_or_create_round'))

    cr_criteria = session.get('cr_criteria', 1.0) # Mặc định là không nhất quán
    candidate_matrices_details = session.get('candidate_pairwise_matrices_details', [])
    
    all_consistent = cr_criteria < 0.1
    inconsistent_candidate_matrices_info = []
    
    criteria_names = session.get('criteria_names', [])
    if not candidate_matrices_details or len(candidate_matrices_details) != len(criteria_names):
        first_missing_idx = 0
        if candidate_matrices_details: # Nếu list không rỗng
            first_missing_idx = next((i for i, detail in enumerate(candidate_matrices_details) if detail is None), len(candidate_matrices_details))
        
        if first_missing_idx < len(criteria_names):
            flash(f"Chưa hoàn thành việc nhập liệu cho ma trận ứng viên của tiêu chí '{criteria_names[first_missing_idx]}'.", "warning")
            return redirect(url_for('input_candidate_comparison_for_criterion', criterion_idx=first_missing_idx))
        else: # Trường hợp lạ
            flash("Dữ liệu ma trận ứng viên không đầy đủ. Vui lòng thử lại.", "danger")
            return redirect(url_for('index_or_create_round'))

    for idx, detail in enumerate(candidate_matrices_details):
        # detail đã được kiểm tra not None ở trên
        if detail['cr'] >= 0.1:
            all_consistent = False
            inconsistent_candidate_matrices_info.append(f"Tiêu chí '{detail.get('criterion_name', criteria_names[idx])}': CR = {detail['cr']:.4f}")
            
    weights_criteria = np.array(session.get('weights_criteria', []))
    candidate_names = session.get('candidate_names', [])
    candidate_ids = session.get('candidate_ids', [])
    
    num_candidates = len(candidate_names)
    num_criteria = len(criteria_names)

    if weights_criteria.shape[0] != num_criteria:
        flash("Lỗi dữ liệu trọng số tiêu chí.", "danger")
        return redirect(url_for('setup_criteria_details'))

    candidate_performance_matrix = np.zeros((num_candidates, num_criteria))
    for crit_idx in range(num_criteria):
        detail = candidate_matrices_details[crit_idx]
        local_weights = np.array(detail['weights'])
        if local_weights.shape[0] == num_candidates:
            candidate_performance_matrix[:, crit_idx] = local_weights
        else: # Lỗi, gán trọng số bằng nhau
            flash(f"Lỗi kích thước trọng số ứng viên cho tiêu chí '{criteria_names[crit_idx]}'. Đã đặt giá trị mặc định.", "warning")
            candidate_performance_matrix[:, crit_idx] = np.ones(num_candidates) / num_candidates
    
    final_scores = candidate_performance_matrix @ weights_criteria
    
    ranked_candidates_list = []
    for i in range(num_candidates):
        ranked_candidates_list.append({
            'id': candidate_ids[i],
            'name': candidate_names[i],
            'score': final_scores[i]
        })
    
    ranked_candidates_display = sorted(ranked_candidates_list, key=lambda x: x['score'], reverse=True)

    round_id = session['round_id']
    if all_consistent:
        CandidateScore.query.filter_by(round_id=round_id).delete()
        for rank_idx, cand_data in enumerate(ranked_candidates_display):
            candidate_score_db = CandidateScore(
                round_id=round_id,
                candidate_id=cand_data['id'],
                total_score=cand_data['score'],
                ranking=rank_idx + 1
            )
            db.session.add(candidate_score_db)
        db.session.commit()
        flash("Kết quả đã được tính toán và lưu trữ thành công!", "success")
    else:
        flash("Một số ma trận không nhất quán. Kết quả chỉ mang tính tham khảo và chưa được lưu vào lịch sử chính thức.", "info")


    # Tạo các biểu đồ
    criterion_weights_image = None
    if weights_criteria.size > 0 and len(criteria_names) == weights_criteria.size:
         criterion_weights_image = create_criterion_weights_chart(criteria_names, weights_criteria)
    
    candidate_score_image = None
    cand_scores_for_chart = [c['score'] for c in ranked_candidates_display]
    cand_names_for_chart = [c['name'] for c in ranked_candidates_display]
    if cand_scores_for_chart and cand_names_for_chart:
        candidate_score_image = create_candidate_scores_chart(cand_scores_for_chart, cand_names_for_chart)
    
    # Tính toán các chỉ số chi tiết cho ma trận tiêu chí
    criteria_matrix = session.get('pairwise_matrix_criteria', [])
    criteria_detailed_metrics = None
    if criteria_matrix:
        try:
            criteria_detailed_metrics = calculate_detailed_consistency_metrics(criteria_matrix)
        except Exception as e:
            print(f"Error calculating criteria detailed metrics: {e}")
    
    # Tính toán các chỉ số chi tiết cho ma trận ứng viên
    candidate_matrices_detailed_metrics = []
    for idx, detail in enumerate(candidate_matrices_details):
        if detail and 'matrix' in detail:
            try:
                detailed_metrics = calculate_detailed_consistency_metrics(detail['matrix'])
                detailed_metrics['criterion_name'] = detail.get('criterion_name', criteria_names[idx] if idx < len(criteria_names) else f'Tiêu chí {idx+1}')
                candidate_matrices_detailed_metrics.append(detailed_metrics)
            except Exception as e:
                print(f"Error calculating detailed metrics for criterion {idx}: {e}")
                candidate_matrices_detailed_metrics.append(None)
        else:
            candidate_matrices_detailed_metrics.append(None)
    
    # Các biểu đồ khác bạn có thể tạo tương tự
    # pairwise_matrix_image = create_pairwise_matrix_visualization(np.array(session.get('pairwise_matrix_criteria', [])), criteria_names)
    # consistency_chart_image = create_consistency_chart(cr_criteria)
    
    # session.pop('ahp_data_active', None) # Ví dụ xóa cờ đang trong quy trình AHP
    
    return render_template('final_results.html',
                           round_name=session.get('round_name'),
                           ranked_candidates=ranked_candidates_display,
                           criteria_names=criteria_names,
                           weights_criteria=weights_criteria.tolist(),
                           cr_criteria=cr_criteria,
                           candidate_names=candidate_names,
                           candidate_matrices_details=candidate_matrices_details,
                           all_consistent=all_consistent,
                           inconsistent_candidate_matrices_info=inconsistent_candidate_matrices_info,
                           pairwise_matrix_criteria=session.get('pairwise_matrix_criteria', []),
                           criterion_weights_image=criterion_weights_image,
                           candidate_score_image=candidate_score_image,
                           criteria_detailed_metrics=criteria_detailed_metrics,
                           candidate_matrices_detailed_metrics=candidate_matrices_detailed_metrics
                           # Thêm các ảnh khác nếu có
                           )

# --- HÀM TẠO BIỂU ĐỒ (GIỮ NGUYÊN HOẶC CHỈNH SỬA NHẸ NẾU CẦN) ---
def create_criterion_weights_chart(criteria, weights):
    if not criteria or not isinstance(weights, (list, np.ndarray)) or not len(weights) or len(criteria) != len(weights):
        return None
    plt.figure(figsize=(10, 6))
    bars = plt.bar(criteria, weights, color='skyblue')
    plt.xlabel('Tiêu chí', fontsize=12)
    plt.ylabel('Trọng số', fontsize=12)
    plt.title('Trọng số các tiêu chí', fontsize=14, fontweight='bold')
    plt.xticks(rotation=45, ha='right', fontsize=10)
    plt.yticks(fontsize=10)
    if len(weights) > 0 and max(weights) > 0 : plt.ylim(0, max(weights) * 1.2)
    else: plt.ylim(0,0.1)


    for bar in bars:
        yval = bar.get_height()
        plt.text(bar.get_x() + bar.get_width()/2, yval + 0.005, f'{yval:.3f}', ha='center', va='bottom', fontsize=9)

    plt.tight_layout()
    img = io.BytesIO()
    plt.savefig(img, format='png')
    img.seek(0)
    encoded_img = base64.b64encode(img.getvalue()).decode('utf-8')
    plt.close()
    return encoded_img

def create_candidate_scores_chart(scores, candidate_names):
    if not candidate_names or not isinstance(scores, (list, np.ndarray)) or not len(scores) or len(candidate_names) != len(scores):
        return None
    plt.figure(figsize=(10, 6))
    bars = plt.bar(candidate_names, scores, color='lightcoral')
    plt.xlabel('Ứng viên', fontsize=12)
    plt.ylabel('Tổng điểm', fontsize=12)
    plt.title('Điểm số tổng của các ứng viên', fontsize=14, fontweight='bold')
    plt.xticks(rotation=45, ha='right', fontsize=10)
    plt.yticks(fontsize=10)
    if len(scores) > 0 and max(scores) > 0: plt.ylim(0, max(scores) * 1.2)
    else: plt.ylim(0,0.1)


    for bar in bars:
        yval = bar.get_height()
        plt.text(bar.get_x() + bar.get_width()/2, yval + 0.005, f'{yval:.3f}', ha='center', va='bottom', fontsize=9)

    plt.tight_layout()
    img = io.BytesIO()
    plt.savefig(img, format='png')
    img.seek(0)
    encoded_img = base64.b64encode(img.getvalue()).decode('utf-8')
    plt.close()
    return encoded_img

# Các hàm create_criteria_comparison_chart, create_raw_data_chart, 
# create_pairwise_matrix_visualization, create_consistency_chart
# bạn có thể giữ lại và gọi chúng trong `calculate_final_ranking` nếu muốn hiển thị các biểu đồ đó.
# Hãy đảm bảo chúng nhận đúng tham số từ dữ liệu AHP mới.

# --- ROUTE CŨ CẦN XEM XÉT KỸ ---
# @app.route('/input/<int:round_id>', methods=['POST'])
# def input_candidates(round_id):
#     # Route này có thể không còn cần thiết nữa nếu bạn hoàn toàn chuyển sang quy trình mới.
#     # Nếu vẫn muốn giữ lại, cần tích hợp rất cẩn thận.
#     pass


# --- CÁC ROUTE QUẢN LÝ LỊCH SỬ, XUẤT PDF ---
# (Cần cập nhật để tương thích với dữ liệu mới, ví dụ: các tiêu chí động,
# ma trận so sánh ứng viên được lưu trong session hoặc DB nếu bạn mở rộng)

# from fpdf2 import FPDF # Đảm bảo import FPDF
import tempfile
import os

@app.route('/export_excel/<int:round_id>')
def export_excel(round_id):
    """Xuất tất cả dữ liệu ra file Excel"""
    try:
        round_info = RecruitmentRound.query.get_or_404(round_id)
        criteria_list_db = RecruitmentCriteria.query.filter_by(round_id=round_id).all()
        criteria_matrix_db = CriteriaMatrix.query.filter_by(round_id=round_id).first()
        candidate_scores_db = CandidateScore.query.filter_by(round_id=round_id).order_by(CandidateScore.ranking).all()
        candidates_db = Candidate.query.filter_by(round_id=round_id).all()

        # Tạo workbook mới
        wb = Workbook()
        
        # Sheet 1: Thông tin đợt tuyển dụng
        ws_info = wb.active
        ws_info.title = "Thông tin đợt tuyển dụng"
        ws_info['A1'] = "Tên đợt tuyển dụng"
        ws_info['B1'] = round_info.round_name
        ws_info['A2'] = "Vị trí"
        ws_info['B2'] = round_info.position
        ws_info['A3'] = "Mô tả"
        ws_info['B3'] = round_info.description or ""
        ws_info['A4'] = "Ngày tạo"
        ws_info['B4'] = round_info.created_at.strftime("%d/%m/%Y %H:%M:%S")        # Sheet 2: Trọng số tiêu chí
        if criteria_matrix_db and criteria_matrix_db.matrix_data:
            ws_criteria = wb.create_sheet("Trọng số tiêu chí")
            matrix_data = criteria_matrix_db.matrix_data
            # Ensure matrix_data is a dictionary (parse if it's a JSON string)
            if isinstance(matrix_data, str):
                matrix_data = json.loads(matrix_data)
            criteria_names = matrix_data.get("criteria_names_at_creation", [c.criterion_name for c in criteria_list_db])
            weights = matrix_data.get("weights", [])
            
            # Headers
            ws_criteria['A1'] = "Tiêu chí"
            ws_criteria['B1'] = "Trọng số"
            
            # Data
            for i, (name, weight) in enumerate(zip(criteria_names, weights), 2):
                ws_criteria[f'A{i}'] = name
                ws_criteria[f'B{i}'] = weight
            
            # Consistency ratio
            ws_criteria[f'A{len(criteria_names)+3}'] = "Chỉ số nhất quán (CR)"
            ws_criteria[f'B{len(criteria_names)+3}'] = criteria_matrix_db.consistency_ratio        # Sheet 3: Ma trận so sánh cặp tiêu chí
        if criteria_matrix_db and criteria_matrix_db.matrix_data:
            ws_matrix = wb.create_sheet("Ma trận so sánh tiêu chí")
            matrix_data = criteria_matrix_db.matrix_data
            # Ensure matrix_data is a dictionary (parse if it's a JSON string)
            if isinstance(matrix_data, str):
                matrix_data = json.loads(matrix_data)
            criteria_names = matrix_data.get("criteria_names_at_creation", [c.criterion_name for c in criteria_list_db])
            pairwise_matrix = matrix_data.get("matrix", [])
            
            # Headers (thêm cột rỗng đầu tiên)
            ws_matrix['A1'] = ""
            for j, name in enumerate(criteria_names, 2):
                ws_matrix.cell(row=1, column=j, value=name)
            
            # Data với tên tiêu chí ở cột đầu
            for i, (name, row_data) in enumerate(zip(criteria_names, pairwise_matrix), 2):
                ws_matrix.cell(row=i, column=1, value=name)
                for j, value in enumerate(row_data, 2):
                    ws_matrix.cell(row=i, column=j, value=value)

        # Sheet 4: Danh sách ứng viên
        ws_candidates = wb.create_sheet("Danh sách ứng viên")
        ws_candidates['A1'] = "STT"
        ws_candidates['B1'] = "Tên ứng viên"
        ws_candidates['C1'] = "Ghi chú"
        
        for i, candidate in enumerate(candidates_db, 2):
            ws_candidates[f'A{i}'] = i - 1
            ws_candidates[f'B{i}'] = candidate.full_name
            ws_candidates[f'C{i}'] = candidate.notes or ""

        # Sheet 5: Kết quả xếp hạng
        if candidate_scores_db:
            ws_ranking = wb.create_sheet("Kết quả xếp hạng")
            ws_ranking['A1'] = "Thứ hạng"
            ws_ranking['B1'] = "Tên ứng viên"
            ws_ranking['C1'] = "Điểm tổng hợp"
            
            for i, score_entry in enumerate(candidate_scores_db, 2):
                candidate_info = Candidate.query.get(score_entry.candidate_id)
                ws_ranking[f'A{i}'] = score_entry.ranking
                ws_ranking[f'B{i}'] = candidate_info.full_name if candidate_info else 'N/A'
                ws_ranking[f'C{i}'] = score_entry.total_score

        # Lấy thêm dữ liệu ma trận so sánh ứng viên từ session nếu có
        if session.get('candidate_pairwise_matrices_details'):
            candidate_matrices_details = session.get('candidate_pairwise_matrices_details', [])
            criteria_names_session = session.get('criteria_names', [])
            candidate_names_session = session.get('candidate_names', [])
            
            for idx, detail in enumerate(candidate_matrices_details):
                if detail is not None:
                    criterion_name = detail.get('criterion_name', criteria_names_session[idx] if idx < len(criteria_names_session) else f'Tiêu chí {idx+1}')
                    matrix = detail.get('matrix', [])
                    cr = detail.get('cr', 0)
                    
                    # Tạo sheet cho mỗi ma trận so sánh ứng viên
                    sheet_name = f"Ma trận UC - {criterion_name}"[:31]  # Excel sheet name limit
                    ws_candidate_matrix = wb.create_sheet(sheet_name)
                    
                    # Headers
                    ws_candidate_matrix['A1'] = ""
                    for j, name in enumerate(candidate_names_session, 2):
                        ws_candidate_matrix.cell(row=1, column=j, value=name)
                    
                    # Data
                    for i, (name, row_data) in enumerate(zip(candidate_names_session, matrix), 2):
                        ws_candidate_matrix.cell(row=i, column=1, value=name)
                        for j, value in enumerate(row_data, 2):
                            ws_candidate_matrix.cell(row=i, column=j, value=value)
                    
                    # Consistency ratio
                    ws_candidate_matrix.cell(row=len(candidate_names_session)+3, column=1, value="Chỉ số nhất quán (CR)")
                    ws_candidate_matrix.cell(row=len(candidate_names_session)+3, column=2, value=cr)        # Lưu file tạm và trả về
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            excel_path = tmp.name
            wb.save(excel_path)
        
        return send_file(
            excel_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'DuLieu_AHP_{round_info.round_name.replace(" ", "_")}.xlsx'
        )
    
    except Exception as e:
        flash(f"Lỗi khi tạo file Excel: {e}", "danger")
        return redirect(url_for('round_detail', round_id=round_id))
    finally:
        if 'excel_path' in locals() and os.path.exists(excel_path):
            try:
                os.unlink(excel_path)
            except Exception as e_unlink:
                app.logger.error(f"Could not remove temp excel file {excel_path}: {e_unlink}")

@app.route('/export_excel_with_charts/<int:round_id>')
def export_excel_with_charts(round_id):
    """Xuất tất cả dữ liệu ra file Excel có kèm biểu đồ"""
    try:
        round_info = RecruitmentRound.query.get_or_404(round_id)
        criteria_list_db = RecruitmentCriteria.query.filter_by(round_id=round_id).all()
        criteria_matrix_db = CriteriaMatrix.query.filter_by(round_id=round_id).first()
        candidate_scores_db = CandidateScore.query.filter_by(round_id=round_id).order_by(CandidateScore.ranking).all()
        candidates_db = Candidate.query.filter_by(round_id=round_id).all()

        # Tạo workbook mới
        wb = Workbook()
        
        # Sheet 1: Thông tin đợt tuyển dụng
        ws_info = wb.active
        ws_info.title = "Thông tin đợt tuyển dụng"
        ws_info['A1'] = "Tên đợt tuyển dụng"
        ws_info['B1'] = round_info.round_name
        ws_info['A2'] = "Vị trí"
        ws_info['B2'] = round_info.position
        ws_info['A3'] = "Mô tả"
        ws_info['B3'] = round_info.description or ""
        ws_info['A4'] = "Ngày tạo"
        ws_info['B4'] = round_info.created_at.strftime("%d/%m/%Y %H:%M:%S")

        # Sheet 2: Trọng số tiêu chí với biểu đồ
        if criteria_matrix_db and criteria_matrix_db.matrix_data:
            ws_criteria = wb.create_sheet("Trọng số tiêu chí")
            matrix_data = criteria_matrix_db.matrix_data
            # Ensure matrix_data is a dictionary (parse if it's a JSON string)
            if isinstance(matrix_data, str):
                matrix_data = json.loads(matrix_data)
            criteria_names = matrix_data.get("criteria_names_at_creation", [c.criterion_name for c in criteria_list_db])
            weights = matrix_data.get("weights", [])
            
            # Headers
            ws_criteria['A1'] = "Tiêu chí"
            ws_criteria['B1'] = "Trọng số"
            
            # Data
            for i, (name, weight) in enumerate(zip(criteria_names, weights), 2):
                ws_criteria[f'A{i}'] = name
                ws_criteria[f'B{i}'] = weight
            
            # Consistency ratio
            ws_criteria[f'A{len(criteria_names)+3}'] = "Chỉ số nhất quán (CR)"
            ws_criteria[f'B{len(criteria_names)+3}'] = criteria_matrix_db.consistency_ratio

            # Tạo biểu đồ cột cho trọng số tiêu chí
            if len(criteria_names) > 0 and len(weights) > 0:
                chart1 = BarChart()
                chart1.type = "col"
                chart1.style = 10
                chart1.title = "Trọng số các tiêu chí"
                chart1.y_axis.title = 'Trọng số'
                chart1.x_axis.title = 'Tiêu chí'

                data = Reference(ws_criteria, min_col=2, min_row=1, max_row=len(criteria_names)+1, max_col=2)
                cats = Reference(ws_criteria, min_col=1, min_row=2, max_row=len(criteria_names)+1)
                chart1.add_data(data, titles_from_data=True)
                chart1.set_categories(cats)
                
                # Đặt biểu đồ vào sheet
                ws_criteria.add_chart(chart1, "D2")

            # Tạo biểu đồ tròn cho trọng số tiêu chí
            if len(criteria_names) > 0 and len(weights) > 0:
                chart2 = PieChart()
                chart2.title = "Phân bổ trọng số tiêu chí"
                
                data = Reference(ws_criteria, min_col=2, min_row=2, max_row=len(criteria_names)+1)
                cats = Reference(ws_criteria, min_col=1, min_row=2, max_row=len(criteria_names)+1)
                chart2.add_data(data, titles_from_data=False)
                chart2.set_categories(cats)
                
                # Đặt biểu đồ vào sheet
                ws_criteria.add_chart(chart2, "D18")

        # Sheet 3: Ma trận so sánh cặp tiêu chí
        if criteria_matrix_db and criteria_matrix_db.matrix_data:
            ws_matrix = wb.create_sheet("Ma trận so sánh tiêu chí")
            matrix_data = criteria_matrix_db.matrix_data
            # Ensure matrix_data is a dictionary (parse if it's a JSON string)
            if isinstance(matrix_data, str):
                matrix_data = json.loads(matrix_data)
            criteria_names = matrix_data.get("criteria_names_at_creation", [c.criterion_name for c in criteria_list_db])
            pairwise_matrix = matrix_data.get("matrix", [])
            
            # Headers (thêm cột rỗng đầu tiên)
            ws_matrix['A1'] = ""
            for j, name in enumerate(criteria_names, 2):
                ws_matrix.cell(row=1, column=j, value=name)
            
            # Data với tên tiêu chí ở cột đầu
            for i, (name, row_data) in enumerate(zip(criteria_names, pairwise_matrix), 2):
                ws_matrix.cell(row=i, column=1, value=name)
                for j, value in enumerate(row_data, 2):
                    ws_matrix.cell(row=i, column=j, value=value)

        # Sheet 4: Danh sách ứng viên
        ws_candidates = wb.create_sheet("Danh sách ứng viên")
        ws_candidates['A1'] = "STT"
        ws_candidates['B1'] = "Tên ứng viên"
        ws_candidates['C1'] = "Ghi chú"
        
        for i, candidate in enumerate(candidates_db, 2):
            ws_candidates[f'A{i}'] = i - 1
            ws_candidates[f'B{i}'] = candidate.full_name
            ws_candidates[f'C{i}'] = candidate.notes or ""

        # Sheet 5: Kết quả xếp hạng
        if candidate_scores_db:
            ws_ranking = wb.create_sheet("Kết quả xếp hạng")
            ws_ranking['A1'] = "Thứ hạng"
            ws_ranking['B1'] = "Tên ứng viên"
            ws_ranking['C1'] = "Điểm tổng hợp"
            
            for i, score_entry in enumerate(candidate_scores_db, 2):
                candidate_info = Candidate.query.get(score_entry.candidate_id)
                ws_ranking[f'A{i}'] = score_entry.ranking
                ws_ranking[f'B{i}'] = candidate_info.full_name if candidate_info else 'N/A'
                ws_ranking[f'C{i}'] = score_entry.total_score

            # Tạo biểu đồ cột cho điểm số ứng viên
            if len(candidate_scores_db) > 0:
                chart3 = BarChart()
                chart3.type = "col"
                chart3.style = 12
                chart3.title = "Điểm tổng hợp các ứng viên"
                chart3.y_axis.title = 'Điểm số'
                chart3.x_axis.title = 'Ứng viên'

                data = Reference(ws_ranking, min_col=3, min_row=1, max_row=len(candidate_scores_db)+1, max_col=3)
                cats = Reference(ws_ranking, min_col=2, min_row=2, max_row=len(candidate_scores_db)+1)
                chart3.add_data(data, titles_from_data=True)
                chart3.set_categories(cats)
                
                # Đặt biểu đồ vào sheet
                ws_ranking.add_chart(chart3, "E2")

        # Lấy thêm dữ liệu ma trận so sánh ứng viên từ session nếu có
        if session.get('candidate_pairwise_matrices_details'):
            candidate_matrices_details = session.get('candidate_pairwise_matrices_details', [])
            criteria_names_session = session.get('criteria_names', [])
            candidate_names_session = session.get('candidate_names', [])
            
            for idx, detail in enumerate(candidate_matrices_details):
                if detail is not None:
                    criterion_name = detail.get('criterion_name', criteria_names_session[idx] if idx < len(criteria_names_session) else f'Tiêu chí {idx+1}')
                    matrix = detail.get('matrix', [])
                    cr = detail.get('cr', 0)
                    
                    # Tạo sheet cho mỗi ma trận so sánh ứng viên
                    sheet_name = f"Ma trận UC - {criterion_name}"[:31]  # Excel sheet name limit
                    ws_candidate_matrix = wb.create_sheet(sheet_name)
                    
                    # Headers
                    ws_candidate_matrix['A1'] = ""
                    for j, name in enumerate(candidate_names_session, 2):
                        ws_candidate_matrix.cell(row=1, column=j, value=name)
                    
                    # Data
                    for i, (name, row_data) in enumerate(zip(candidate_names_session, matrix), 2):
                        ws_candidate_matrix.cell(row=i, column=1, value=name)
                        for j, value in enumerate(row_data, 2):
                            ws_candidate_matrix.cell(row=i, column=j, value=value)
                    
                    # Consistency ratio
                    ws_candidate_matrix.cell(row=len(candidate_names_session)+3, column=1, value="Chỉ số nhất quán (CR)")
                    ws_candidate_matrix.cell(row=len(candidate_names_session)+3, column=2, value=cr)

        # Lưu file tạm và trả về
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            excel_path = tmp.name
            wb.save(excel_path)
        
        return send_file(
            excel_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'DuLieu_AHP_BieuDo_{round_info.round_name.replace(" ", "_")}.xlsx'
        )
    
    except Exception as e:
        flash(f"Lỗi khi tạo file Excel với biểu đồ: {e}", "danger")
        return redirect(url_for('round_detail', round_id=round_id))
    finally:
        if 'excel_path' in locals() and os.path.exists(excel_path):
            try:
                os.unlink(excel_path)
            except Exception as e_unlink:
                app.logger.error(f"Could not remove temp excel file {excel_path}: {e_unlink}")

@app.route('/export_report/<int:round_id>')
def export_report(round_id):
    round_info = RecruitmentRound.query.get_or_404(round_id)
    criteria_list_db = RecruitmentCriteria.query.filter_by(round_id=round_id).all() # Lấy từ DB
    criteria_matrix_db = CriteriaMatrix.query.filter_by(round_id=round_id).first()
    candidate_scores_db = CandidateScore.query.filter_by(round_id=round_id).order_by(CandidateScore.ranking).all()

    pdf = FPDF()
    pdf.add_page()
    try: # Thêm font hỗ trợ Unicode
        pdf.add_font('DejaVu', '', 'DejaVuSans.ttf', uni=True)
        pdf.set_font('DejaVu', '', 16)
    except RuntimeError: # Fallback nếu font không tìm thấy
        pdf.set_font('Arial', 'B', 16)
        flash("Font DejaVu Sans không tìm thấy, báo cáo PDF có thể không hiển thị đúng tiếng Việt.", "warning")


    pdf.cell(0, 10, f'Báo Cáo Tuyển Dụng: {round_info.round_name}', 0, 1, 'C')
    pdf.set_font('DejaVu' if 'DejaVu' in pdf.font_family else 'Arial', '', 12)
    pdf.cell(0, 10, f'Vị trí: {round_info.position}', 0, 1)
    if round_info.description:
        pdf.multi_cell(0, 10, f'Mô tả: {round_info.description}') # Dùng multi_cell cho text dài
    pdf.cell(0, 10, f'Ngày tạo: {round_info.created_at.strftime("%d/%m/%Y %H:%M:%S")}', 0, 1)
    pdf.ln(5)    # Trọng số và CR của Tiêu chí
    if criteria_matrix_db and criteria_matrix_db.matrix_data:
        matrix_data = criteria_matrix_db.matrix_data # Đây là dict từ JSON
        # Ensure matrix_data is a dictionary (parse if it's a JSON string)
        if isinstance(matrix_data, str):
            matrix_data = json.loads(matrix_data)
        criteria_names_from_matrix = matrix_data.get("criteria_names_at_creation", [c.criterion_name for c in criteria_list_db])
        weights_criteria = matrix_data.get("weights", [])
        
        pdf.set_font('DejaVu' if 'DejaVu' in pdf.font_family else 'Arial', 'B', 12)
        pdf.cell(0, 10, '1. Phân Tích Tiêu Chí', 0, 1)
        pdf.set_font('DejaVu' if 'DejaVu' in pdf.font_family else 'Arial', '', 11)
        pdf.cell(0, 8, f'Chỉ số nhất quán (CR) ma trận tiêu chí: {criteria_matrix_db.consistency_ratio:.4f}', 0, 1)
        
        pdf.cell(0, 8, 'Trọng số các tiêu chí:', 0, 1)
        for i, name in enumerate(criteria_names_from_matrix):
            if i < len(weights_criteria):
                pdf.cell(0, 7, f'  - {name}: {weights_criteria[i]:.4f}', 0, 1)
        pdf.ln(5)
    
    # Kết quả Xếp Hạng Ứng Viên
    if candidate_scores_db:
        pdf.set_font('DejaVu' if 'DejaVu' in pdf.font_family else 'Arial', 'B', 12)
        pdf.cell(0, 10, '2. Kết Quả Xếp Hạng Ứng Viên', 0, 1)
        pdf.set_font('DejaVu' if 'DejaVu' in pdf.font_family else 'Arial', '', 11)
        
        # Header bảng
        pdf.set_fill_color(230, 230, 230)
        pdf.cell(15, 7, 'Hạng', 1, 0, 'C', 1)
        pdf.cell(100, 7, 'Tên Ứng Viên', 1, 0, 'C', 1)
        pdf.cell(65, 7, 'Điểm Tổng Hợp AHP', 1, 1, 'C', 1)

        for score_entry in candidate_scores_db:
            candidate_info = Candidate.query.get(score_entry.candidate_id)
            pdf.cell(15, 7, str(score_entry.ranking), 1, 0, 'C')
            pdf.cell(100, 7, candidate_info.full_name if candidate_info else 'N/A', 1, 0, 'L')
            pdf.cell(65, 7, f'{score_entry.total_score:.4f}', 1, 1, 'R')
        pdf.ln(5)

    # (Tùy chọn) Thêm chi tiết ma trận so sánh ứng viên nếu bạn có lưu chúng vào DB
    # Hoặc nếu chúng đang có trong session khi export được gọi từ trang kết quả
    # Điều này sẽ phức tạp hơn và cần truy cập session hoặc query DB mới.

    # Xuất file PDF
    # File path can be tricky with tempfile on some systems, ensure permissions.
    try:
        # Create a temporary file that is automatically deleted after sending.
        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp:
            pdf_path = tmp.name
            pdf.output(pdf_path)
        
        return send_file(
            pdf_path,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'BaoCao_{round_info.round_name.replace(" ", "_")}.pdf'
            # Không dùng max_age=0 nếu muốn trình duyệt cache
        )
    except Exception as e:
        flash(f"Lỗi khi tạo file PDF: {e}", "danger")
        return redirect(url_for('round_detail', round_id=round_id)) # Hoặc một trang lỗi
    finally:
        if 'pdf_path' in locals() and os.path.exists(pdf_path): # Đảm bảo dọn dẹp file tạm
            try:
                os.unlink(pdf_path)
            except Exception as e_unlink:
                app.logger.error(f"Could not remove temp pdf file {pdf_path}: {e_unlink}")


@app.route('/history')
def history():
    try:
        rounds = RecruitmentRound.query.order_by(RecruitmentRound.created_at.desc()).all()
    except Exception as e:
        flash(f"Không thể tải lịch sử: {e}", "danger")
        rounds = []
    return render_template('history.html', rounds=rounds)

@app.route('/round/<int:round_id>')
def round_detail(round_id):
    round_info = RecruitmentRound.query.get_or_404(round_id)
    criteria_list_db = RecruitmentCriteria.query.filter_by(round_id=round_id).order_by(RecruitmentCriteria.criterion_id).all() # Sắp xếp để nhất quán
    criteria_matrix_db = CriteriaMatrix.query.filter_by(round_id=round_id).first()
    candidates_db = Candidate.query.filter_by(round_id=round_id).all()
    scores_db = CandidateScore.query.filter_by(round_id=round_id).order_by(CandidateScore.ranking).all()

    matrix_display_data = None
    weights_display = None
    criteria_names_display = [c.criterion_name for c in criteria_list_db] # Lấy tên từ DB làm chuẩn

    if criteria_matrix_db and criteria_matrix_db.matrix_data:
        try:
            # matrix_data từ DB là một dict
            matrix_display_data = criteria_matrix_db.matrix_data.get("matrix")
            weights_display = criteria_matrix_db.matrix_data.get("weights")
            # Có thể kiểm tra criteria_names_at_creation nếu muốn so sánh
        except Exception as e: # Xử lý nếu matrix_data không phải JSON hợp lệ (dù không nên)
            flash(f"Lỗi đọc dữ liệu ma trận: {e}", "warning")
            matrix_display_data = [] # Hoặc giá trị mặc định khác
            weights_display = []

    ranked_display = []
    for score_entry in scores_db:
        candidate_info = Candidate.query.get(score_entry.candidate_id)
        if candidate_info:
            ranked_display.append({
                'name': candidate_info.full_name, 
                'score': score_entry.total_score,
                'rank': score_entry.ranking
            })
    
    # Nếu muốn hiển thị cả các ma trận so sánh ứng viên, bạn cần query từ DB (nếu đã lưu)
    # hoặc thiết kế lại cách lưu/truy xuất. Hiện tại, chúng không được lưu vào DB trong logic đã cung cấp.

    return render_template('round_detail.html', 
                           round=round_info, 
                           criteria_names=criteria_names_display, 
                           matrix_data=matrix_display_data, # Chỉ truyền phần ma trận
                           weights_criteria=weights_display, # Truyền trọng số
                           consistency_ratio_criteria=criteria_matrix_db.consistency_ratio if criteria_matrix_db else None,
                           candidates=candidates_db, 
                           ranked_scores=ranked_display)


@app.route('/delete_round/<int:round_id>')
def delete_round(round_id):
    try:
        # Đảm bảo xóa theo đúng thứ tự khóa ngoại
        CandidateScore.query.filter_by(round_id=round_id).delete()
        # Nếu có bảng CandidateCriterionMatrices, xóa ở đây
        CriteriaMatrix.query.filter_by(round_id=round_id).delete()
        Candidate.query.filter_by(round_id=round_id).delete()
        RecruitmentCriteria.query.filter_by(round_id=round_id).delete()
        RecruitmentRound.query.filter_by(round_id=round_id).delete()
        db.session.commit()
        flash("Đợt tuyển dụng và tất cả dữ liệu liên quan đã được xóa.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Lỗi khi xóa đợt tuyển dụng: {str(e)}", "danger")
    return redirect(url_for('history'))

# API endpoint cho tính toán consistency real-time
@app.route('/api/calculate_criteria_consistency', methods=['POST'])
def api_calculate_criteria_consistency():
    """API endpoint để tính toán consistency cho ma trận tiêu chí"""
    try:
        data = request.get_json()
        
        if not data or 'matrix' not in data:
            return jsonify({
                'success': False,
                'error': 'Matrix data không hợp lệ'
            }), 400
        
        matrix_data = data['matrix']
        n = len(matrix_data)
        
        # Validate matrix structure
        if not all(len(row) == n for row in matrix_data):
            return jsonify({
                'success': False,
                'error': 'Ma trận không vuông'
            }), 400
        
        # Convert to numpy array
        matrix = np.array(matrix_data, dtype=float)
        
        # Calculate detailed consistency metrics
        metrics = calculate_detailed_consistency_metrics(matrix)
        
        # Calculate priority vector
        priority_vector = calculate_priority_vector(matrix)
        
        return jsonify({
            'success': True,
            'lambda_max': float(metrics['lambda_max']),
            'ci': float(metrics['ci']),
            'cr': float(metrics['cr']),
            'ri': float(metrics['ri']),
            'priority_vector': priority_vector.tolist() if hasattr(priority_vector, 'tolist') else list(priority_vector),
            'is_consistent': float(metrics['cr']) < 0.1
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Lỗi tính toán: {str(e)}'
        }), 500

@app.route('/api/calculate_candidate_consistency', methods=['POST'])
def api_calculate_candidate_consistency():
    """API endpoint để tính toán consistency cho ma trận ứng viên"""
    try:
        data = request.get_json()
        
        if not data or 'matrix' not in data:
            return jsonify({
                'success': False,
                'error': 'Matrix data không hợp lệ'
            }), 400
        
        matrix_data = data['matrix']
        n = len(matrix_data)
        
        # Validate matrix structure
        if not all(len(row) == n for row in matrix_data):
            return jsonify({
                'success': False,
                'error': 'Ma trận không vuông'
            }), 400
        
        # Convert to numpy array
        matrix = np.array(matrix_data, dtype=float)
        
        # Calculate detailed consistency metrics
        metrics = calculate_detailed_consistency_metrics(matrix)
        
        # Calculate priority vector
        priority_vector = calculate_priority_vector(matrix)
        
        return jsonify({
            'success': True,
            'lambda_max': float(metrics['lambda_max']),
            'ci': float(metrics['ci']),
            'cr': float(metrics['cr']),
            'ri': float(metrics['ri']),
            'priority_vector': priority_vector.tolist() if hasattr(priority_vector, 'tolist') else list(priority_vector),
            'is_consistent': float(metrics['cr']) < 0.1
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Lỗi tính toán: {str(e)}'
        }), 500

@app.route('/export_excel_single_sheet/<int:round_id>')
def export_excel_single_sheet(round_id):
    """Xuất tất cả dữ liệu ra file Excel trong 1 sheet duy nhất"""
    try:
        round_info = RecruitmentRound.query.get_or_404(round_id)
        criteria_list_db = RecruitmentCriteria.query.filter_by(round_id=round_id).all()
        criteria_matrix_db = CriteriaMatrix.query.filter_by(round_id=round_id).first()
        candidate_scores_db = CandidateScore.query.filter_by(round_id=round_id).order_by(CandidateScore.ranking).all()
        candidates_db = Candidate.query.filter_by(round_id=round_id).all()

        # Tạo workbook với 1 sheet duy nhất
        wb = Workbook()
        ws = wb.active
        ws.title = "Báo cáo AHP tổng hợp"
        
        current_row = 1
        
        # Phần 1: Thông tin đợt tuyển dụng
        ws[f'A{current_row}'] = "=== THÔNG TIN ĐỢT TUYỂN DỤNG ==="
        ws[f'A{current_row}'].font = ws[f'A{current_row}'].font.copy(bold=True, size=14)
        current_row += 1
        
        ws[f'A{current_row}'] = "Tên đợt:"
        ws[f'B{current_row}'] = round_info.round_name
        current_row += 1
        
        ws[f'A{current_row}'] = "Vị trí:"
        ws[f'B{current_row}'] = round_info.position
        current_row += 1
        
        ws[f'A{current_row}'] = "Mô tả:"
        ws[f'B{current_row}'] = round_info.description or ""
        current_row += 1
        
        ws[f'A{current_row}'] = "Ngày tạo:"
        ws[f'B{current_row}'] = round_info.created_at.strftime("%d/%m/%Y %H:%M:%S")
        current_row += 3
        
        # Phần 2: Kết quả xếp hạng (phần quan trọng nhất)
        if candidate_scores_db:
            ws[f'A{current_row}'] = "=== KẾT QUẢ XẾP HẠNG ==="
            ws[f'A{current_row}'].font = ws[f'A{current_row}'].font.copy(bold=True, size=14)
            current_row += 1
            
            # Headers
            ws[f'A{current_row}'] = "Thứ hạng"
            ws[f'B{current_row}'] = "Tên ứng viên"
            ws[f'C{current_row}'] = "Điểm tổng hợp"
            
            # Make header bold
            for col in ['A', 'B', 'C']:
                ws[f'{col}{current_row}'].font = ws[f'{col}{current_row}'].font.copy(bold=True)
            current_row += 1
            
            # Data
            for score_entry in candidate_scores_db:
                candidate_info = Candidate.query.get(score_entry.candidate_id)
                ws[f'A{current_row}'] = score_entry.ranking
                ws[f'B{current_row}'] = candidate_info.full_name if candidate_info else 'N/A'
                ws[f'C{current_row}'] = round(score_entry.total_score, 4)
                current_row += 1
            
            current_row += 2
        
        # Phần 3: Trọng số tiêu chí
        if criteria_matrix_db and criteria_matrix_db.matrix_data:
            ws[f'A{current_row}'] = "=== TRỌNG SỐ TIÊU CHÍ ==="
            ws[f'A{current_row}'].font = ws[f'A{current_row}'].font.copy(bold=True, size=14)
            current_row += 1
            
            matrix_data = criteria_matrix_db.matrix_data
            if isinstance(matrix_data, str):
                matrix_data = json.loads(matrix_data)
            criteria_names = matrix_data.get("criteria_names_at_creation", [c.criterion_name for c in criteria_list_db])
            weights = matrix_data.get("weights", [])
            
            # Headers
            ws[f'A{current_row}'] = "Tiêu chí"
            ws[f'B{current_row}'] = "Trọng số"
            ws[f'C{current_row}'] = "Tỷ lệ %"
            
            # Make header bold
            for col in ['A', 'B', 'C']:
                ws[f'{col}{current_row}'].font = ws[f'{col}{current_row}'].font.copy(bold=True)
            current_row += 1
            
            # Data
            for name, weight in zip(criteria_names, weights):
                ws[f'A{current_row}'] = name
                ws[f'B{current_row}'] = round(weight, 4)
                ws[f'C{current_row}'] = f"{round(weight * 100, 2)}%"
                current_row += 1
            
            # Consistency ratio
            current_row += 1
            ws[f'A{current_row}'] = "Chỉ số nhất quán (CR):"
            ws[f'B{current_row}'] = round(criteria_matrix_db.consistency_ratio, 4)
            ws[f'C{current_row}'] = "✓ Chấp nhận" if criteria_matrix_db.consistency_ratio < 0.1 else "⚠ Cần xem xét"
            current_row += 3
        
        # Phần 4: Danh sách tất cả ứng viên
        if candidates_db:
            ws[f'A{current_row}'] = "=== DANH SÁCH ỨNG VIÊN ==="
            ws[f'A{current_row}'].font = ws[f'A{current_row}'].font.copy(bold=True, size=14)
            current_row += 1
            
            # Headers
            ws[f'A{current_row}'] = "STT"
            ws[f'B{current_row}'] = "Tên ứng viên"
            ws[f'C{current_row}'] = "Ghi chú"
            
            # Make header bold
            for col in ['A', 'B', 'C']:
                ws[f'{col}{current_row}'].font = ws[f'{col}{current_row}'].font.copy(bold=True)
            current_row += 1
            
            # Data
            for i, candidate in enumerate(candidates_db, 1):
                ws[f'A{current_row}'] = i
                ws[f'B{current_row}'] = candidate.full_name
                ws[f'C{current_row}'] = candidate.notes or ""
                current_row += 1
            
            current_row += 2
        
        # Phần 5: Ma trận so sánh tiêu chí (tóm tắt)
        if criteria_matrix_db and criteria_matrix_db.matrix_data:
            ws[f'A{current_row}'] = "=== MA TRẬN SO SÁNH TIÊU CHÍ ==="
            ws[f'A{current_row}'].font = ws[f'A{current_row}'].font.copy(bold=True, size=14)
            current_row += 1
            
            matrix_data = criteria_matrix_db.matrix_data
            if isinstance(matrix_data, str):
                matrix_data = json.loads(matrix_data)
            criteria_names = matrix_data.get("criteria_names_at_creation", [c.criterion_name for c in criteria_list_db])
            pairwise_matrix = matrix_data.get("matrix", [])
            
            if criteria_names and pairwise_matrix:
                # Headers với tên tiêu chí
                ws[f'A{current_row}'] = ""
                for j, name in enumerate(criteria_names):
                    col_letter = chr(66 + j)  # B, C, D, ...
                    ws[f'{col_letter}{current_row}'] = name
                    ws[f'{col_letter}{current_row}'].font = ws[f'{col_letter}{current_row}'].font.copy(bold=True)
                current_row += 1
                
                # Data với tên tiêu chí ở cột đầu
                for i, (name, row_data) in enumerate(zip(criteria_names, pairwise_matrix), 2):
                    ws[f'A{current_row}'] = name
                    ws[f'A{current_row}'].font = ws[f'A{current_row}'].font.copy(bold=True)
                    for j, value in enumerate(row_data):
                        col_letter = chr(66 + j)  # B, C, D, ...
                        ws[f'{col_letter}{current_row}'] = round(value, 3)
                    current_row += 1
                
                current_row += 2
        
        # Phần 6: Tóm tắt ma trận so sánh ứng viên (chỉ hiển thị trọng số cuối)
        if session.get('candidate_pairwise_matrices_details'):
            ws[f'A{current_row}'] = "=== TRỌNG SỐ ỨNG VIÊN THEO TỪNG TIÊU CHÍ ==="
            ws[f'A{current_row}'].font = ws[f'A{current_row}'].font.copy(bold=True, size=14)
            current_row += 1
            
            candidate_matrices_details = session.get('candidate_pairwise_matrices_details', [])
            criteria_names_session = session.get('criteria_names', [])
            candidate_names_session = session.get('candidate_names', [])
            
            if candidate_names_session:
                # Headers
                ws[f'A{current_row}'] = "Tiêu chí"
                for i, candidate_name in enumerate(candidate_names_session):
                    col_letter = chr(66 + i)  # B, C, D, ...
                    ws[f'{col_letter}{current_row}'] = candidate_name
                    ws[f'{col_letter}{current_row}'].font = ws[f'{col_letter}{current_row}'].font.copy(bold=True)
                ws[f'A{current_row}'].font = ws[f'A{current_row}'].font.copy(bold=True)
                current_row += 1
                
                # Data - trọng số của từng ứng viên theo từng tiêu chí
                for idx, detail in enumerate(candidate_matrices_details):
                    if detail is not None:
                        criterion_name = detail.get('criterion_name', criteria_names_session[idx] if idx < len(criteria_names_session) else f'Tiêu chí {idx+1}')
                        weights_detail = detail.get('weights', [])
                        
                        ws[f'A{current_row}'] = criterion_name
                        for i, weight in enumerate(weights_detail):
                            col_letter = chr(66 + i)  # B, C, D, ...
                            ws[f'{col_letter}{current_row}'] = round(weight, 4)
                        current_row += 1
                
                current_row += 2
        
        # Điều chỉnh độ rộng cột
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        
        # Lưu file tạm và trả về
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            excel_path = tmp.name
            wb.save(excel_path)
        
        return send_file(
            excel_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'BaoCao_AHP_TongHop_{round_info.round_name.replace(" ", "_")}.xlsx'
        )
    
    except Exception as e:
        flash(f"Lỗi khi tạo file Excel tổng hợp: {e}", "danger")
        return redirect(url_for('round_detail', round_id=round_id))
    finally:
        if 'excel_path' in locals() and os.path.exists(excel_path):
            try:
                os.unlink(excel_path)
            except Exception as e_unlink:
                app.logger.error(f"Could not remove temp excel file {excel_path}: {e_unlink}")

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True)
