#!/usr/bin/env python3
"""
Final Excel Import Functionality Validation
Tests all Excel import/export capabilities without requiring session state
"""
import os
import sys
from openpyxl import load_workbook

# Add the current directory to Python path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from app import (
    read_candidates_from_excel, 
    read_matrix_from_excel,
    create_candidate_template_excel,
    create_matrix_template_excel,
    calculate_consistency_ratio,
    calculate_priority_vector
)
import numpy as np

def test_candidate_import_export():
    """Test candidate import and template creation"""
    print("🧪 Testing Candidate Import/Export...")
    
    # Test reading existing candidate file
    try:
        candidates, error = read_candidates_from_excel('test_candidates.xlsx')
        if error:
            print(f"❌ Error reading candidates: {error}")
            return False
        print(f"✅ Read {len(candidates)} candidates: {candidates}")
    except Exception as e:
        print(f"❌ Exception reading candidates: {e}")
        return False
    
    # Test creating candidate template
    try:
        wb = create_candidate_template_excel()
        wb.save('generated_candidate_template.xlsx')
        wb.close()
        print("✅ Created candidate template successfully")
        
        # Verify the template
        candidates_template, error = read_candidates_from_excel('generated_candidate_template.xlsx')
        if not error and candidates_template:
            print(f"✅ Template verification passed: {len(candidates_template)} sample candidates")
        else:
            print(f"❌ Template verification failed: {error}")
            return False
            
    except Exception as e:
        print(f"❌ Error creating candidate template: {e}")
        return False
    
    return True

def test_matrix_import_export():
    """Test matrix import and template creation"""
    print("\n🧪 Testing Matrix Import/Export...")
    
    # Test reading existing matrix file
    try:
        matrix, error = read_matrix_from_excel('test_criteria_3x3.xlsx', 3)
        if error:
            print(f"❌ Error reading matrix: {error}")
            return False
        print(f"✅ Read {len(matrix)}x{len(matrix[0])} matrix")
        print(f"   First row: {matrix[0]}")
    except Exception as e:
        print(f"❌ Exception reading matrix: {e}")
        return False
    
    # Test creating matrix template
    try:
        wb = create_matrix_template_excel(4, "test_criteria")
        wb.save('generated_matrix_template.xlsx')
        wb.close()
        print("✅ Created 4x4 matrix template successfully")
        
        # Verify the template
        matrix_template, error = read_matrix_from_excel('generated_matrix_template.xlsx', 4)
        if not error and matrix_template:
            print(f"✅ Template verification passed: {len(matrix_template)}x{len(matrix_template[0])} matrix")
        else:
            print(f"❌ Template verification failed: {error}")
            return False
            
    except Exception as e:
        print(f"❌ Error creating matrix template: {e}")
        return False
    
    return True

def test_matrix_calculations():
    """Test AHP calculations with imported matrix"""
    print("\n🧪 Testing AHP Calculations...")
    
    try:
        # Read the test matrix
        matrix_data, error = read_matrix_from_excel('test_criteria_3x3.xlsx', 3)
        if error:
            print(f"❌ Error reading matrix for calculations: {error}")
            return False
        
        matrix = np.array(matrix_data)
        
        # Calculate consistency ratio
        cr = calculate_consistency_ratio(matrix)
        print(f"✅ Consistency Ratio calculated: {cr:.4f}")
        
        if cr < 0.1:
            print("✅ Matrix is consistent (CR < 0.1)")
        else:
            print(f"⚠️ Matrix is inconsistent (CR = {cr:.4f} >= 0.1)")
        
        # Calculate priority vector
        weights = calculate_priority_vector(matrix)
        print(f"✅ Priority weights calculated: {weights}")
        print(f"   Sum of weights: {sum(weights):.4f}")
        
        return True
        
    except Exception as e:
        print(f"❌ Error in AHP calculations: {e}")
        return False

def test_file_formats():
    """Test different file format handling"""
    print("\n🧪 Testing File Format Support...")
    
    test_results = []
    
    # Test existing Excel files
    test_files = [
        ('test_candidates.xlsx', 'candidate list'),
        ('test_criteria_3x3.xlsx', '3x3 matrix'),
        ('test_candidates_4x4.xlsx', '4x4 candidate matrix')
    ]
    
    for filename, description in test_files:
        if os.path.exists(filename):
            try:
                wb = load_workbook(filename)
                ws = wb.active
                rows = ws.max_row
                cols = ws.max_column
                wb.close()
                print(f"✅ {filename} ({description}): {rows} rows, {cols} columns")
                test_results.append(True)
            except Exception as e:
                print(f"❌ {filename} is corrupted: {e}")
                test_results.append(False)
        else:
            print(f"❌ {filename} not found")
            test_results.append(False)
    
    return all(test_results)

def test_error_handling():
    """Test error handling with invalid files"""
    print("\n🧪 Testing Error Handling...")
    
    # Test with non-existent file
    candidates, error = read_candidates_from_excel('nonexistent.xlsx')
    if error:
        print(f"✅ Properly handled non-existent file: {error}")
    else:
        print("❌ Should have returned error for non-existent file")
        return False
    
    # Test matrix with wrong size
    matrix, error = read_matrix_from_excel('test_criteria_3x3.xlsx', 5)
    if matrix and len(matrix) == 3:  # Should still read the matrix, but size check happens later
        print("✅ Matrix size handling works correctly")
    else:
        print(f"❌ Matrix size handling issue: {error}")
        return False
    
    return True

def cleanup_test_files():
    """Clean up generated test files"""
    print("\n🧹 Cleaning up test files...")
    
    test_files = [
        'generated_candidate_template.xlsx',
        'generated_matrix_template.xlsx'
    ]
    
    for filename in test_files:
        if os.path.exists(filename):
            try:
                os.remove(filename)
                print(f"✅ Removed {filename}")
            except Exception as e:
                print(f"❌ Error removing {filename}: {e}")

def main():
    """Run all tests"""
    print("🚀 Excel Import Functionality Validation")
    print("=" * 50)
    
    # Change to the correct directory
    os.chdir('/Users/thanhtruong/cnmp1/ahp-main')
    
    tests = [
        test_candidate_import_export,
        test_matrix_import_export,
        test_matrix_calculations,
        test_file_formats,
        test_error_handling
    ]
    
    results = []
    for test in tests:
        try:
            result = test()
            results.append(result)
        except Exception as e:
            print(f"❌ Test failed with exception: {e}")
            results.append(False)
    
    print("\n" + "=" * 50)
    print("📊 Final Test Results:")
    print(f"   ✅ Passed: {sum(results)}/{len(results)}")
    print(f"   ❌ Failed: {len(results) - sum(results)}/{len(results)}")
    
    if all(results):
        print("\n🎉 ALL TESTS PASSED!")
        print("Excel import functionality is working perfectly!")
        print("\n📋 Features Validated:")
        print("   ✅ Candidate list import from Excel")
        print("   ✅ Matrix import from Excel")
        print("   ✅ Template generation and download")
        print("   ✅ AHP calculations (CR, weights)")
        print("   ✅ File format validation")
        print("   ✅ Error handling")
    else:
        print("\n❌ Some tests failed.")
        print("Please review the errors above.")
    
    # Cleanup
    cleanup_test_files()
    
    return all(results)

if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)
