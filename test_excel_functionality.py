#!/usr/bin/env python3
"""
Comprehensive test suite for Excel import functionality
"""
import os
import requests
import tempfile
from openpyxl import load_workbook

BASE_URL = "http://127.0.0.1:5000"

def test_candidate_template_download():
    """Test downloading candidate template"""
    print("🧪 Testing candidate template download...")
    try:
        response = requests.get(f"{BASE_URL}/download_candidate_template")
        if response.status_code == 200:
            # Save and verify the downloaded file
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                tmp.write(response.content)
                tmp_path = tmp.name
            
            # Verify it's a valid Excel file
            wb = load_workbook(tmp_path)
            ws = wb.active
            header = ws['A1'].value
            sample_data = ws['A2'].value
            wb.close()
            os.unlink(tmp_path)
            
            print(f"✅ Template downloaded successfully")
            print(f"   Header: {header}")
            print(f"   Sample: {sample_data}")
            return True
        else:
            print(f"❌ Failed to download template: {response.status_code}")
            return False
    except Exception as e:
        print(f"❌ Error: {e}")
        return False

def test_files_exist():
    """Test that all required test files exist"""
    print("🧪 Testing test files existence...")
    required_files = [
        'test_candidates.xlsx',
        'test_criteria_3x3.xlsx', 
        'test_candidates_4x4.xlsx'
    ]
    
    all_exist = True
    for file in required_files:
        if os.path.exists(file):
            print(f"✅ {file} exists")
        else:
            print(f"❌ {file} missing")
            all_exist = False
    
    return all_exist

def test_excel_file_content():
    """Test content of Excel files"""
    print("🧪 Testing Excel file content...")
    
    # Test candidate file
    try:
        wb = load_workbook('test_candidates.xlsx')
        ws = wb.active
        candidates = []
        for row in range(2, ws.max_row + 1):
            if ws[f'A{row}'].value:
                candidates.append(ws[f'A{row}'].value)
        wb.close()
        print(f"✅ test_candidates.xlsx contains {len(candidates)} candidates: {candidates}")
    except Exception as e:
        print(f"❌ Error reading test_candidates.xlsx: {e}")
        return False
    
    # Test criteria matrix file
    try:
        wb = load_workbook('test_criteria_3x3.xlsx')
        ws = wb.active
        matrix = []
        for row in range(1, 4):
            matrix_row = []
            for col in range(1, 4):
                matrix_row.append(ws.cell(row=row, column=col).value)
            matrix.append(matrix_row)
        wb.close()
        print(f"✅ test_criteria_3x3.xlsx contains valid 3x3 matrix")
        print(f"   First row: {matrix[0]}")
    except Exception as e:
        print(f"❌ Error reading test_criteria_3x3.xlsx: {e}")
        return False
    
    return True

def main():
    """Run all tests"""
    print("🚀 Starting Excel Import Functionality Tests")
    print("=" * 50)
    
    # Change to the correct directory
    os.chdir('/Users/thanhtruong/cnmp1/ahp-main')
    
    tests = [
        test_files_exist,
        test_excel_file_content,
        test_candidate_template_download
    ]
    
    results = []
    for test in tests:
        print()
        result = test()
        results.append(result)
    
    print("\n" + "=" * 50)
    print("📊 Test Results:")
    print(f"   Passed: {sum(results)}/{len(results)}")
    print(f"   Failed: {len(results) - sum(results)}/{len(results)}")
    
    if all(results):
        print("🎉 All tests passed! Excel import functionality is working correctly.")
    else:
        print("❌ Some tests failed. Please check the errors above.")

if __name__ == "__main__":
    main()
